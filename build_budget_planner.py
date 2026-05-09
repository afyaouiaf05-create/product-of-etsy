"""
Budget_Planner_2026_All_In_One.xlsx generator
Commercial-quality digital product for Etsy
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
import copy

# ── Colour Palette ──────────────────────────────────────────────────────────
C_PRIMARY   = "2E5AAC"   # deep blue
C_SECONDARY = "4A90D9"   # medium blue
C_ACCENT    = "F4A261"   # warm orange
C_SUCCESS   = "2A9D8F"   # teal green
C_DANGER    = "E76F51"   # coral red
C_WARNING   = "E9C46A"   # yellow
C_DARK      = "264653"   # dark teal
C_LIGHT     = "F8F9FA"   # off-white
C_INPUT     = "FFFEF0"   # light yellow
C_WHITE     = "FFFFFF"
C_BORDER    = "DEE2E6"
C_GREEN_BG  = "D4EDDA"   # income green background
C_RED_BG    = "F8D7DA"   # expense red background
C_GRAY      = "6C757D"
C_LIGHT_GRAY= "E9ECEF"

# ── Reusable style factories ─────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(name="Calibri", size=11, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color=C_BORDER)
    n = None
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

def medium_border(sides="all"):
    s = Side(style="medium", color=C_DARK)
    return Border(top=s, bottom=s, left=s, right=s)

def double_border_bottom():
    return Border(
        top=Side(style="thin", color=C_BORDER),
        bottom=Side(style="double", color=C_DARK),
        left=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
    )

def header_style(ws, row, col, value, bg=C_PRIMARY, fg=C_WHITE,
                 size=12, bold=True, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(size=size, bold=bold, color=fg)
    c.alignment = align(h=h_align, v="center")
    c.border = thin_border()
    return c

def data_cell(ws, row, col, value=None, num_format=None,
              bg=C_WHITE, bold=False, h_align="left", wrap=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(size=11, bold=bold, color=color)
    c.alignment = align(h=h_align, v="center", wrap=wrap)
    c.border = thin_border()
    if num_format:
        c.number_format = num_format
    return c

def input_cell(ws, row, col, value=None, num_format=None, h_align="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(C_INPUT)
    c.font = font(size=11)
    c.alignment = align(h=h_align, v="center")
    c.border = thin_border()
    if num_format:
        c.number_format = num_format
    return c

def currency_fmt(ws, row, col, value=None, bg=C_WHITE, bold=False):
    return data_cell(ws, row, col, value, num_format='"$"#,##0.00',
                     bg=bg, bold=bold, h_align="right")

def pct_fmt(ws, row, col, value=None, bg=C_WHITE, bold=False):
    return data_cell(ws, row, col, value, num_format='0.0%',
                     bg=bg, bold=bold, h_align="right")

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def merge_and_style(ws, r1, c1, r2, c2, value, bg=C_PRIMARY, fg=C_WHITE,
                    size=14, bold=True, h_align="center", v_align="center", wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.fill = fill(bg)
    c.font = font(size=size, bold=bold, color=fg)
    c.alignment = align(h=h_align, v=v_align, wrap=wrap)
    return c

# ── Sheet 1 : Start Here ─────────────────────────────────────────────────────

def build_start_here(wb):
    ws = wb.create_sheet("Start Here")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_PRIMARY

    # Set column widths
    widths = {"A": 3, "B": 28, "C": 28, "D": 28, "E": 28, "F": 3}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    # Row heights
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 18
    for r in range(5, 60):
        ws.row_dimensions[r].height = 22

    # ── Hero banner ──────────────────────────────────────────────────────
    merge_and_style(ws, 2, 2, 2, 5,
                    "2026 All-In-One Automated Budget Planner",
                    bg=C_PRIMARY, fg=C_WHITE, size=24, bold=True)

    merge_and_style(ws, 3, 2, 3, 5,
                    "✦  Just enter your transactions, everything else calculates automatically  ✦",
                    bg=C_DARK, fg=C_WARNING, size=13, bold=False)

    merge_and_style(ws, 4, 2, 4, 5,
                    "Personal Finance Made Simple — No Excel Knowledge Required",
                    bg=C_SECONDARY, fg=C_WHITE, size=11, bold=False)

    # ── How It Works ─────────────────────────────────────────────────────
    row = 6
    merge_and_style(ws, row, 2, row, 5,
                    "HOW IT WORKS", bg=C_ACCENT, fg=C_WHITE, size=14, bold=True)
    row += 1

    steps = [
        ("STEP 1 — SETUP", "Open 'Category Setup' sheet.\nSet your monthly income and budget targets for each spending category.", C_SUCCESS),
        ("STEP 2 — TRACK", "Open 'Monthly Tracker' sheet.\nEnter each transaction as it happens. Just type the date, category, and amount.", C_SECONDARY),
        ("STEP 3 — REVIEW", "Open 'Dashboard' sheet.\nSee your spending analysis, 50/30/20 breakdown, and savings progress instantly.", C_PRIMARY),
    ]
    for title, desc, color in steps:
        ws.row_dimensions[row].height = 18
        merge_and_style(ws, row, 2, row, 2, title, bg=color, fg=C_WHITE, size=11, bold=True, h_align="center")
        ws.row_dimensions[row + 1].height = 40
        merge_and_style(ws, row + 1, 2, row + 1, 2, desc, bg=C_LIGHT, fg="333333", size=10, bold=False, wrap=True)
        row += 2

    # Position step boxes side by side (redo — merge across cols properly)
    # Reset and redo steps in columns B, C, D
    # (rows 7–12 were used — clear and redo)
    for r in range(7, 15):
        ws.row_dimensions[r].height = 50

    # Step titles row 7
    row = 7
    step_data = [
        ("① SETUP\nOpen 'Category Setup'\nSet income & budget targets", C_SUCCESS),
        ("② TRACK\nOpen 'Monthly Tracker'\nEnter each transaction", C_SECONDARY),
        ("③ REVIEW\nOpen 'Dashboard'\nSee instant analysis", C_PRIMARY),
    ]
    cols = [2, 3, 4]
    for (title, color), col in zip(step_data, cols):
        ws.merge_cells(start_row=7, start_column=col, end_row=9, end_column=col)
        c = ws.cell(row=7, column=col, value=title)
        c.fill = fill(color)
        c.font = font(size=11, bold=True, color=C_WHITE)
        c.alignment = align(h="center", v="center", wrap=True)

    # Connector column E row 7-9
    merge_and_style(ws, 7, 5, 9, 5, "→ Repeat\nEach Month", bg=C_ACCENT, fg=C_WHITE, size=10, bold=True)

    # ── Feature List ─────────────────────────────────────────────────────
    row = 11
    merge_and_style(ws, row, 2, row, 5,
                    "✦  WHAT'S INCLUDED", bg=C_DARK, fg=C_WHITE, size=13, bold=True)
    row += 1

    features = [
        ("✔  Auto-Calculations", "All totals and balances update instantly as you type."),
        ("✔  Color Coding",      "Green for income, red for expenses — at a glance clarity."),
        ("✔  50/30/20 Analysis", "Built-in method to split Needs / Wants / Savings automatically."),
        ("✔  Progress Bars",     "Visual savings goal bars built with spreadsheet formulas."),
        ("✔  Charts & KPIs",     "Dashboard cards showing income, expenses, balance & savings rate."),
        ("✔  Annual Overview",   "12-month bird's-eye view of every income and expense category."),
        ("✔  Debt Tracker",      "Track balances, interest rates, and months to payoff."),
        ("✔  Savings Goals",     "6 preset goals with progress tracking — add your own easily."),
    ]
    for feat, desc in features:
        ws.row_dimensions[row].height = 22
        c1 = data_cell(ws, row, 2, feat, bg=C_LIGHT, bold=True, color=C_DARK)
        c2 = data_cell(ws, row, 3, desc, bg=C_WHITE, wrap=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        row += 1

    # ── Tips ─────────────────────────────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 2, row, 5, "PRO TIPS FOR BEST RESULTS",
                    bg=C_WARNING, fg=C_DARK, size=12, bold=True)
    row += 1
    tips = [
        "💡  Update your tracker weekly so no transactions are forgotten.",
        "💡  Use the same category names consistently for accurate reports.",
        "💡  Review your Dashboard at the end of each month to spot trends.",
        "💡  Set realistic budget targets in Category Setup — be honest!",
    ]
    for tip in tips:
        ws.row_dimensions[row].height = 22
        merge_and_style(ws, row, 2, row, 5, tip, bg=C_LIGHT, fg="333333", size=10, bold=False, h_align="left")
        row += 1

    # ── Sheet Navigation ─────────────────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 2, row, 5, "SHEET NAVIGATION",
                    bg=C_SECONDARY, fg=C_WHITE, size=12, bold=True)
    row += 1
    nav = [
        ("📋 Category Setup",  "Set your monthly income & expense budgets"),
        ("💳 Monthly Tracker", "Log every income & expense transaction"),
        ("📊 Dashboard",       "Visual KPIs, 50/30/20 analysis, budget vs actual"),
        ("🎯 Savings Goals",   "Track progress toward your savings targets"),
        ("💸 Debt Tracker",    "Monitor balances and payoff timelines"),
        ("📅 Annual Overview", "Full-year income & expense summary"),
        ("🗓  Monthly Calendar","Bill due dates and income dates for January 2026"),
    ]
    for sheet, desc in nav:
        ws.row_dimensions[row].height = 22
        data_cell(ws, row, 2, sheet, bg=C_LIGHT, bold=True, color=C_PRIMARY)
        c = data_cell(ws, row, 3, desc, bg=C_WHITE)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        row += 1

    # ── Footer ───────────────────────────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 2, row, 5,
                    "Compatible with Google Sheets & Excel  |  Digital Download  |  Personal Use",
                    bg=C_DARK, fg=C_LIGHT, size=10, bold=False)


# ── Sheet 2 : Category Setup ─────────────────────────────────────────────────

def build_category_setup(wb):
    ws = wb.create_sheet("Category Setup")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_SUCCESS

    widths = {"A": 3, "B": 26, "C": 16, "D": 18, "E": 3}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 22
    for r in range(4, 60):
        ws.row_dimensions[r].height = 22

    # Title
    merge_and_style(ws, 2, 2, 2, 4, "CATEGORY SETUP", bg=C_PRIMARY, fg=C_WHITE, size=18, bold=True)
    merge_and_style(ws, 3, 2, 3, 4,
                    "Set your budget targets below — yellow cells are yours to edit",
                    bg=C_SECONDARY, fg=C_WHITE, size=11)

    # ── EXPENSE CATEGORIES ───────────────────────────────────────────────
    row = 5
    merge_and_style(ws, row, 2, row, 4, "EXPENSE CATEGORIES", bg=C_DARK, fg=C_WHITE, size=12, bold=True)
    row += 1
    for col, hdr in zip([2, 3, 4], ["Category", "Budget Type (Need/Want/Savings)", "Monthly Target ($)"]):
        header_style(ws, row, col, hdr, bg=C_SECONDARY, size=11)
    row += 1

    expense_cats = [
        ("Housing",         "Need",    2000.00),
        ("Utilities",       "Need",     180.00),
        ("Groceries",       "Need",     500.00),
        ("Transportation",  "Need",     350.00),
        ("Insurance",       "Need",     250.00),
        ("Healthcare",      "Need",     150.00),
        ("Dining Out",      "Want",     200.00),
        ("Entertainment",   "Want",     100.00),
        ("Shopping",        "Want",     150.00),
        ("Travel",          "Want",     100.00),
        ("Hobbies",         "Want",      80.00),
        ("Subscriptions",   "Want",      60.00),
        ("Emergency Fund",  "Savings",  300.00),
        ("Retirement",      "Savings",  400.00),
        ("Investments",     "Savings",  200.00),
        ("Debt Payment",    "Need",     500.00),
        ("Education",       "Want",     100.00),
        ("Gifts/Donations", "Want",      80.00),
    ]

    expense_start = row
    type_colors = {"Need": C_DANGER, "Want": C_ACCENT, "Savings": C_SUCCESS}

    for name, btype, target in expense_cats:
        data_cell(ws, row, 2, name, bg=C_WHITE)
        tc = data_cell(ws, row, 3, btype, bg=C_WHITE, h_align="center",
                       color=type_colors.get(btype, "000000"), bold=True)
        input_cell(ws, row, 4, target, num_format='"$"#,##0.00')
        row += 1

    expense_end = row - 1

    # Total Budget Target
    row += 1
    data_cell(ws, row, 2, "TOTAL BUDGET TARGET", bg=C_LIGHT, bold=True)
    data_cell(ws, row, 3, "", bg=C_LIGHT)
    c = ws.cell(row=row, column=4,
                value=f"=SUM(D{expense_start}:D{expense_end})")
    c.fill = fill(C_PRIMARY)
    c.font = font(size=12, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = thin_border()
    c.number_format = '"$"#,##0.00'
    budget_total_row = row

    # ── INCOME CATEGORIES ────────────────────────────────────────────────
    row += 2
    merge_and_style(ws, row, 2, row, 4, "INCOME CATEGORIES", bg=C_DARK, fg=C_WHITE, size=12, bold=True)
    row += 1
    for col, hdr in zip([2, 3, 4], ["Income Source", "Type", "Monthly Expected ($)"]):
        header_style(ws, row, col, hdr, bg=C_SUCCESS, size=11)
    row += 1

    income_cats = [
        ("Salary",      "Primary",   5500.00),
        ("Side Hustle", "Secondary",  500.00),
        ("Investments", "Passive",    200.00),
        ("Freelance",   "Variable",   300.00),
        ("Other",       "Variable",   100.00),
    ]

    income_start = row
    for name, itype, amt in income_cats:
        data_cell(ws, row, 2, name, bg=C_WHITE)
        data_cell(ws, row, 3, itype, bg=C_WHITE, h_align="center")
        input_cell(ws, row, 4, amt, num_format='"$"#,##0.00')
        row += 1
    income_end = row - 1

    # Total Expected Income
    row += 1
    data_cell(ws, row, 2, "TOTAL EXPECTED INCOME", bg=C_LIGHT, bold=True)
    data_cell(ws, row, 3, "", bg=C_LIGHT)
    c = ws.cell(row=row, column=4,
                value=f"=SUM(D{income_start}:D{income_end})")
    c.fill = fill(C_SUCCESS)
    c.font = font(size=12, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = thin_border()
    c.number_format = '"$"#,##0.00'
    income_total_row = row

    # Projected Surplus
    row += 1
    data_cell(ws, row, 2, "PROJECTED MONTHLY SURPLUS", bg=C_LIGHT, bold=True)
    data_cell(ws, row, 3, "", bg=C_LIGHT)
    c = ws.cell(row=row, column=4,
                value=f"=D{income_total_row}-D{budget_total_row}")
    c.fill = fill(C_ACCENT)
    c.font = font(size=12, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = double_border_bottom()
    c.number_format = '"$"#,##0.00'

    # Legend
    row += 2
    merge_and_style(ws, row, 2, row, 4, "COLOR LEGEND", bg=C_LIGHT_GRAY, fg=C_DARK, size=11, bold=True)
    row += 1
    legend_items = [
        (C_DANGER,  "Need — essential expenses (target: ~50% of income)"),
        (C_ACCENT,  "Want — lifestyle spending (target: ~30% of income)"),
        (C_SUCCESS, "Savings — future & goals (target: ~20% of income)"),
        (C_INPUT,   "Yellow cells = your data to edit"),
    ]
    for color, label in legend_items:
        c = ws.cell(row=row, column=2, value="  ■")
        c.fill = fill(color)
        c.font = font(size=12, bold=True, color=color)
        c.alignment = align(h="center", v="center")
        c.border = thin_border()
        merge_and_style(ws, row, 3, row, 4, label, bg=C_WHITE, fg="333333",
                        size=10, bold=False, h_align="left")
        row += 1


# ── Sheet 3 : Monthly Tracker ─────────────────────────────────────────────────

def build_monthly_tracker(wb):
    ws = wb.create_sheet("Monthly Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ACCENT
    ws.freeze_panes = "A9"

    col_widths = {
        "A": 12, "B": 10, "C": 18, "D": 26, "E": 14,
        "F": 16, "G": 10, "H": 22,
    }
    for col, w in col_widths.items():
        set_col_width(ws, col, w)

    # Title rows
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8
    for r in range(5, 250):
        ws.row_dimensions[r].height = 20

    merge_and_style(ws, 2, 1, 2, 8, "MONTHLY TRANSACTION TRACKER",
                    bg=C_ACCENT, fg=C_WHITE, size=18, bold=True)
    merge_and_style(ws, 3, 1, 3, 8,
                    "Enter each income or expense below — totals calculate automatically",
                    bg=C_DARK, fg=C_WHITE, size=11)

    # Column headers
    hdr_row = 5
    headers = ["Date", "Type\n(Income/Expense)", "Category", "Description",
               "Amount ($)", "Payment Method", "Month", "Notes"]
    ws.row_dimensions[hdr_row].height = 32
    for i, h in enumerate(headers, 1):
        header_style(ws, hdr_row, i, h, bg=C_PRIMARY, size=11)
        ws.cell(row=hdr_row, column=i).alignment = align(h="center", v="center", wrap=True)

    # Sample transactions
    samples = [
        ("2026-01-05", "Income",  "Salary",         "January Paycheck",          5500.00, "Direct Deposit", "January", ""),
        ("2026-01-05", "Expense", "Housing",         "Rent — January",            2000.00, "Bank Transfer",  "January", "Monthly rent"),
        ("2026-01-05", "Expense", "Insurance",       "Renters Insurance",          125.00, "Auto-Pay",       "January", ""),
        ("2026-01-07", "Expense", "Groceries",       "Whole Foods weekly shop",    127.43, "Debit Card",     "January", ""),
        ("2026-01-08", "Expense", "Utilities",       "Electric bill",               94.20, "Auto-Pay",       "January", ""),
        ("2026-01-10", "Expense", "Transportation",  "Subway monthly pass",         132.00, "Credit Card",   "January", ""),
        ("2026-01-12", "Expense", "Dining Out",      "Dinner with friends",          68.50, "Credit Card",   "January", ""),
        ("2026-01-14", "Expense", "Subscriptions",   "Netflix + Spotify",            26.98, "Credit Card",   "January", ""),
        ("2026-01-15", "Income",  "Freelance",       "Website design project",      350.00, "PayPal",        "January", ""),
        ("2026-01-16", "Expense", "Groceries",       "Trader Joe's",                 89.65, "Debit Card",    "January", ""),
        ("2026-01-18", "Expense", "Entertainment",   "Movie tickets x2",             28.00, "Credit Card",   "January", ""),
        ("2026-01-20", "Expense", "Emergency Fund",  "Monthly savings transfer",    300.00, "Bank Transfer", "January", "Automated"),
        ("2026-01-22", "Expense", "Healthcare",      "Prescription refill",          42.00, "Debit Card",    "January", ""),
        ("2026-01-25", "Expense", "Shopping",        "Amazon — household items",     56.99, "Credit Card",   "January", ""),
        ("2026-01-30", "Expense", "Debt Payment",    "Student loan payment",        350.00, "Auto-Pay",      "January", ""),
    ]

    data_start = hdr_row + 1
    for i, row_data in enumerate(samples):
        r = data_start + i
        date_val, type_val, cat, desc, amt, pmt, month, notes = row_data

        # Type-based background
        row_bg = C_GREEN_BG if type_val == "Income" else C_RED_BG

        data_cell(ws, r, 1, date_val, bg=row_bg, h_align="center")
        tc = data_cell(ws, r, 2, type_val, bg=row_bg, bold=True, h_align="center",
                       color=C_SUCCESS if type_val == "Income" else C_DANGER)
        data_cell(ws, r, 3, cat,  bg=row_bg)
        data_cell(ws, r, 4, desc, bg=row_bg, wrap=True)
        c = ws.cell(row=r, column=5, value=amt)
        c.fill = fill(row_bg)
        c.font = font(size=11, bold=True,
                      color=C_SUCCESS if type_val == "Income" else C_DANGER)
        c.alignment = align(h="right", v="center")
        c.border = thin_border()
        c.number_format = '"$"#,##0.00'
        data_cell(ws, r, 6, pmt,   bg=row_bg, h_align="center")
        data_cell(ws, r, 7, month, bg=row_bg, h_align="center")
        data_cell(ws, r, 8, notes, bg=row_bg, wrap=True)

    # Empty rows 200 more
    empty_start = data_start + len(samples)
    empty_end = empty_start + 199
    for r in range(empty_start, empty_end + 1):
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.fill = fill(C_WHITE if r % 2 == 0 else C_LIGHT)
            c.border = thin_border()
            c.font = font(size=11)
            if col == 5:
                c.number_format = '"$"#,##0.00'
                c.alignment = align(h="right", v="center")
            elif col in [1, 6, 7]:
                c.alignment = align(h="center", v="center")
            else:
                c.alignment = align(h="left", v="center", wrap=(col == 4))

    # ── TOTALS ───────────────────────────────────────────────────────────
    total_row = empty_end + 2
    ws.row_dimensions[total_row].height = 28
    ws.row_dimensions[total_row + 1].height = 28
    ws.row_dimensions[total_row + 2].height = 32

    merge_and_style(ws, total_row, 1, total_row, 4,
                    "TOTAL INCOME", bg=C_SUCCESS, fg=C_WHITE, size=12, bold=True, h_align="right")
    c = ws.cell(row=total_row, column=5,
                value=f'=SUMIF(B{data_start}:B{empty_end},"Income",E{data_start}:E{empty_end})')
    c.fill = fill(C_SUCCESS)
    c.font = font(size=13, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = medium_border()
    c.number_format = '"$"#,##0.00'
    income_total_cell = f"E{total_row}"

    merge_and_style(ws, total_row + 1, 1, total_row + 1, 4,
                    "TOTAL EXPENSES", bg=C_DANGER, fg=C_WHITE, size=12, bold=True, h_align="right")
    c = ws.cell(row=total_row + 1, column=5,
                value=f'=SUMIF(B{data_start}:B{empty_end},"Expense",E{data_start}:E{empty_end})')
    c.fill = fill(C_DANGER)
    c.font = font(size=13, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = medium_border()
    c.number_format = '"$"#,##0.00'
    expense_total_cell = f"E{total_row+1}"

    merge_and_style(ws, total_row + 2, 1, total_row + 2, 4,
                    "NET BALANCE (Income − Expenses)", bg=C_PRIMARY, fg=C_WHITE,
                    size=13, bold=True, h_align="right")
    c = ws.cell(row=total_row + 2, column=5,
                value=f"={income_total_cell}-{expense_total_cell}")
    c.fill = fill(C_PRIMARY)
    c.font = font(size=14, bold=True, color=C_WHITE)
    c.alignment = align(h="right", v="center")
    c.border = double_border_bottom()
    c.number_format = '"$"#,##0.00'

    # Store important row numbers for cross-sheet references
    ws["A1"] = data_start        # hidden helper — data start row
    ws["A1"].font = font(size=1, color=C_WHITE)

    return {
        "data_start": data_start,
        "data_end": empty_end,
        "income_cell": income_total_cell,
        "expense_cell": expense_total_cell,
    }


# ── Sheet 4 : Dashboard ───────────────────────────────────────────────────────

def build_dashboard(wb, tracker_info):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_SECONDARY

    widths = {"A": 3, "B": 22, "C": 16, "D": 16, "E": 14, "F": 16, "G": 3}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    for r in range(1, 80):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 22

    # Title
    merge_and_style(ws, 2, 2, 2, 6, "FINANCIAL DASHBOARD — 2026",
                    bg=C_PRIMARY, fg=C_WHITE, size=20, bold=True)
    merge_and_style(ws, 3, 2, 3, 6,
                    "Live summary — updates automatically as you enter transactions",
                    bg=C_SECONDARY, fg=C_WHITE, size=11)

    ds = tracker_info["data_start"]
    de = tracker_info["data_end"]

    # ── KPI CARDS ────────────────────────────────────────────────────────
    row = 5
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 42
    ws.row_dimensions[row + 2].height = 22
    ws.row_dimensions[row + 3].height = 8

    kpi_configs = [
        ("TOTAL INCOME",   f"=SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\",'Monthly Tracker'!E{ds}:E{de})",   C_SUCCESS, "💰"),
        ("TOTAL EXPENSES", f"=SUMIF('Monthly Tracker'!B{ds}:B{de},\"Expense\",'Monthly Tracker'!E{ds}:E{de})",  C_DANGER,  "💳"),
        ("NET BALANCE",    f"=SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\",'Monthly Tracker'!E{ds}:E{de})-SUMIF('Monthly Tracker'!B{ds}:B{de},\"Expense\",'Monthly Tracker'!E{ds}:E{de})", C_PRIMARY,   "📊"),
        ("SAVINGS RATE",   None, C_ACCENT, "🎯"),
    ]
    kpi_cols = [2, 3, 4, 5]

    for (label, formula, color, icon), col in zip(kpi_configs, kpi_cols):
        # Label row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)
        lc = ws.cell(row=row, column=col, value=f"{icon} {label}")
        lc.fill = fill(color)
        lc.font = font(size=10, bold=True, color=C_WHITE)
        lc.alignment = align(h="center", v="center")

        # Value row
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col)
        vc = ws.cell(row=row + 1, column=col)
        vc.fill = fill(color)
        vc.font = font(size=18, bold=True, color=C_WHITE)
        vc.alignment = align(h="center", v="center")
        if label == "SAVINGS RATE":
            vc.value = (
                f"=IF(SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\",'Monthly Tracker'!E{ds}:E{de})=0,0,"
                f"(SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\",'Monthly Tracker'!E{ds}:E{de})"
                f"-SUMIF('Monthly Tracker'!B{ds}:B{de},\"Expense\",'Monthly Tracker'!E{ds}:E{de}))"
                f"/SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\",'Monthly Tracker'!E{ds}:E{de}))"
            )
            vc.number_format = '0.0%'
        else:
            vc.value = formula
            vc.number_format = '"$"#,##0.00'

        # Bottom accent bar
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col)
        bc = ws.cell(row=row + 2, column=col, value="")
        bc.fill = fill(C_DARK)

    # ── BUDGET vs ACTUAL TABLE ────────────────────────────────────────────
    row = 10
    ws.row_dimensions[row].height = 28
    merge_and_style(ws, row, 2, row, 6, "BUDGET vs ACTUAL — CURRENT MONTH",
                    bg=C_DARK, fg=C_WHITE, size=13, bold=True)
    row += 1

    tbl_headers = ["Category", "Budget Target", "Actual Spent", "Variance", "Status"]
    tbl_cols = [2, 3, 4, 5, 6]
    ws.row_dimensions[row].height = 26
    for h, col in zip(tbl_headers, tbl_cols):
        header_style(ws, row, col, h, bg=C_SECONDARY, size=11)
    row += 1

    bva_cats = [
        ("Housing",         2000.00),
        ("Groceries",        500.00),
        ("Transportation",   350.00),
        ("Utilities",        180.00),
        ("Dining Out",       200.00),
        ("Entertainment",    100.00),
        ("Shopping",         150.00),
        ("Emergency Fund",   300.00),
        ("Debt Payment",     500.00),
        ("Healthcare",       150.00),
    ]

    for i, (cat, budget) in enumerate(bva_cats):
        bg = C_WHITE if i % 2 == 0 else C_LIGHT
        ws.row_dimensions[row].height = 22
        data_cell(ws, row, 2, cat, bg=bg, bold=False)

        # Budget target
        bc = ws.cell(row=row, column=3, value=budget)
        bc.fill = fill(bg)
        bc.font = font(size=11)
        bc.alignment = align(h="right", v="center")
        bc.border = thin_border()
        bc.number_format = '"$"#,##0.00'

        # Actual spent (SUMIFS from tracker)
        ac = ws.cell(
            row=row, column=4,
            value=(f'=SUMIFS(\'Monthly Tracker\'!E{ds}:E{de},'
                   f'\'Monthly Tracker\'!B{ds}:B{de},"Expense",'
                   f'\'Monthly Tracker\'!C{ds}:C{de},"{cat}")')
        )
        ac.fill = fill(bg)
        ac.font = font(size=11)
        ac.alignment = align(h="right", v="center")
        ac.border = thin_border()
        ac.number_format = '"$"#,##0.00'

        # Variance
        vc = ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        vc.fill = fill(bg)
        vc.font = font(size=11)
        vc.alignment = align(h="right", v="center")
        vc.border = thin_border()
        vc.number_format = '"$"#,##0.00'

        # Status
        sc = ws.cell(row=row, column=6,
                     value=f'=IF(D{row}<=C{row},"✓ Under Budget","✗ Over Budget")')
        sc.fill = fill(bg)
        sc.font = font(size=11)
        sc.alignment = align(h="center", v="center")
        sc.border = thin_border()
        row += 1

    # Conditional formatting for status column
    status_range = f"F12:F{row-1}"
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'F12="✓ Under Budget"'],
        fill=fill(C_GREEN_BG),
        font=font(bold=True, color=C_SUCCESS)
    ))
    ws.conditional_formatting.add(status_range, FormulaRule(
        formula=[f'F12="✗ Over Budget"'],
        fill=fill(C_RED_BG),
        font=font(bold=True, color=C_DANGER)
    ))

    # ── 50/30/20 ANALYSIS ────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 28
    merge_and_style(ws, row, 2, row, 6, "50/30/20 BUDGET ANALYSIS",
                    bg=C_DARK, fg=C_WHITE, size=13, bold=True)
    row += 1

    analysis_headers = ["Category", "Target %", "Target Amount", "Actual Amount", "Status"]
    for h, col in zip(analysis_headers, tbl_cols):
        header_style(ws, row, col, h, bg=C_ACCENT, fg=C_WHITE, size=11)
    row += 1

    income_formula = (f"SUMIF('Monthly Tracker'!B{ds}:B{de},\"Income\","
                      f"'Monthly Tracker'!E{ds}:E{de})")
    needs_cats  = ["Housing","Utilities","Groceries","Transportation","Insurance",
                   "Healthcare","Debt Payment"]
    wants_cats  = ["Dining Out","Entertainment","Shopping","Travel","Hobbies",
                   "Subscriptions","Education","Gifts/Donations"]
    savings_cats= ["Emergency Fund","Retirement","Investments"]

    def sumifs_for_cats(cats):
        parts = "+".join(
            f'SUMIFS(\'Monthly Tracker\'!E{ds}:E{de},'
            f'\'Monthly Tracker\'!B{ds}:B{de},"Expense",'
            f'\'Monthly Tracker\'!C{ds}:C{de},"{c}")'
            for c in cats
        )
        return parts

    analysis_rows = [
        ("Needs (50%)",   0.50, sumifs_for_cats(needs_cats),   C_DANGER),
        ("Wants (30%)",   0.30, sumifs_for_cats(wants_cats),   C_ACCENT),
        ("Savings (20%)", 0.20, sumifs_for_cats(savings_cats), C_SUCCESS),
    ]

    for label, target_pct, actual_formula, color in analysis_rows:
        ws.row_dimensions[row].height = 24
        data_cell(ws, row, 2, label, bg=C_LIGHT, bold=True, color=color)

        tc = ws.cell(row=row, column=3, value=target_pct)
        tc.fill = fill(C_LIGHT)
        tc.font = font(size=11, bold=True)
        tc.alignment = align(h="center", v="center")
        tc.border = thin_border()
        tc.number_format = '0%'

        tamt = ws.cell(row=row, column=4, value=f"={income_formula}*{target_pct}")
        tamt.fill = fill(C_LIGHT)
        tamt.font = font(size=11)
        tamt.alignment = align(h="right", v="center")
        tamt.border = thin_border()
        tamt.number_format = '"$"#,##0.00'

        aamt = ws.cell(row=row, column=5, value=f"={actual_formula}")
        aamt.fill = fill(C_LIGHT)
        aamt.font = font(size=11)
        aamt.alignment = align(h="right", v="center")
        aamt.border = thin_border()
        aamt.number_format = '"$"#,##0.00'

        stat = ws.cell(row=row, column=6,
                       value=f'=IF(E{row}<=D{row},"✓ On Track","⚠ Adjust")')
        stat.fill = fill(C_LIGHT)
        stat.font = font(size=11, bold=True)
        stat.alignment = align(h="center", v="center")
        stat.border = thin_border()
        row += 1

    # ── QUICK TIPS ────────────────────────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 2, row, 6,
                    "💡  Tip: Add transactions to 'Monthly Tracker' and this dashboard updates automatically!",
                    bg=C_WARNING, fg=C_DARK, size=10, bold=False, h_align="left")


# ── Sheet 5 : Savings Goals ───────────────────────────────────────────────────

def build_savings_goals(wb):
    ws = wb.create_sheet("Savings Goals")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_SUCCESS

    widths = {"A": 3, "B": 22, "C": 16, "D": 16, "E": 14, "F": 12, "G": 26, "H": 3}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    for r in range(1, 50):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 22

    # Title
    merge_and_style(ws, 2, 2, 2, 7, "SAVINGS GOALS TRACKER",
                    bg=C_SUCCESS, fg=C_WHITE, size=18, bold=True)
    merge_and_style(ws, 3, 2, 3, 7,
                    "Track progress toward your financial goals — yellow cells are yours to edit",
                    bg=C_DARK, fg=C_WHITE, size=11)

    # Headers
    row = 5
    headers = ["Goal Name", "Target Amount", "Current Saved", "Remaining",
               "Progress %", "Visual Progress ████░░░░"]
    for i, h in enumerate(headers):
        header_style(ws, row, i + 2, h, bg=C_PRIMARY, size=11)
    row += 1

    goals = [
        ("🚨 Emergency Fund",   10000.00,  3200.00),
        ("✈️  Vacation Fund",    3000.00,  1750.00),
        ("🚗 New Car",          15000.00,  4500.00),
        ("🏠 Home Down Payment",50000.00,  8000.00),
        ("💍 Wedding",          20000.00,  6500.00),
        ("📚 Education Fund",    8000.00,  2400.00),
    ]

    goal_start = row
    for name, target, saved in goals:
        ws.row_dimensions[row].height = 24
        data_cell(ws, row, 2, name, bg=C_WHITE, bold=True)

        tc = ws.cell(row=row, column=3, value=target)
        tc.fill = fill(C_LIGHT)
        tc.font = font(size=11)
        tc.alignment = align(h="right", v="center")
        tc.border = thin_border()
        tc.number_format = '"$"#,##0.00'

        sc = input_cell(ws, row, 4, saved, num_format='"$"#,##0.00')

        rc = ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        rc.fill = fill(C_WHITE)
        rc.font = font(size=11)
        rc.alignment = align(h="right", v="center")
        rc.border = thin_border()
        rc.number_format = '"$"#,##0.00'

        pc = ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
        pc.fill = fill(C_WHITE)
        pc.font = font(size=11, bold=True)
        pc.alignment = align(h="right", v="center")
        pc.border = thin_border()
        pc.number_format = '0.0%'

        # REPT progress bar (20 chars wide)
        bar_c = ws.cell(
            row=row, column=7,
            value=(f'=IFERROR(REPT("█",ROUND(F{row}*20,0))&REPT("░",20-ROUND(F{row}*20,0)),"░░░░░░░░░░░░░░░░░░░░")')
        )
        bar_c.fill = fill(C_WHITE)
        bar_c.font = font(size=11, color=C_SUCCESS)
        bar_c.alignment = align(h="left", v="center")
        bar_c.border = thin_border()
        row += 1

    # Colour scale on progress %
    ws.conditional_formatting.add(
        f"F{goal_start}:F{row-1}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="E76F51",
            mid_type="num", mid_value=0.5, mid_color="E9C46A",
            end_type="num", end_value=1, end_color="2A9D8F",
        )
    )

    # Empty user rows
    row += 1
    merge_and_style(ws, row, 2, row, 7, "ADD YOUR OWN GOALS BELOW",
                    bg=C_LIGHT_GRAY, fg=C_GRAY, size=11, bold=True)
    row += 1
    for _ in range(8):
        ws.row_dimensions[row].height = 24
        data_cell(ws, row, 2, "", bg=C_INPUT)

        tc = input_cell(ws, row, 3, None, num_format='"$"#,##0.00')
        sc = input_cell(ws, row, 4, None, num_format='"$"#,##0.00')

        rc = ws.cell(row=row, column=5, value=f"=IF(C{row}=0,\"\",C{row}-D{row})")
        rc.fill = fill(C_WHITE)
        rc.font = font(size=11)
        rc.alignment = align(h="right")
        rc.border = thin_border()
        rc.number_format = '"$"#,##0.00'

        pc = ws.cell(row=row, column=6,
                     value=f"=IF(C{row}=0,\"\",IF(D{row}=0,0,D{row}/C{row}))")
        pc.fill = fill(C_WHITE)
        pc.font = font(size=11)
        pc.alignment = align(h="right")
        pc.border = thin_border()
        pc.number_format = '0.0%'

        bar_c = ws.cell(
            row=row, column=7,
            value=(f'=IF(C{row}=0,"",IFERROR(REPT("█",ROUND(F{row}*20,0))&REPT("░",20-ROUND(F{row}*20,0)),""))')
        )
        bar_c.fill = fill(C_WHITE)
        bar_c.font = font(size=11, color=C_SUCCESS)
        bar_c.alignment = align(h="left")
        bar_c.border = thin_border()
        row += 1

    # ── SUMMARY ───────────────────────────────────────────────────────────
    row += 1
    merge_and_style(ws, row, 2, row, 7, "TOTALS SUMMARY",
                    bg=C_DARK, fg=C_WHITE, size=12, bold=True)
    row += 1
    summ = [
        ("Total Goal Amount",   f"=SUM(C{goal_start}:C{goal_start+5})"),
        ("Total Saved",         f"=SUM(D{goal_start}:D{goal_start+5})"),
        ("Total Remaining",     f"=SUM(E{goal_start}:E{goal_start+5})"),
    ]
    for label, formula in summ:
        ws.row_dimensions[row].height = 24
        data_cell(ws, row, 2, label, bg=C_LIGHT, bold=True)
        for col in range(3, 7):
            ws.cell(row=row, column=col).fill = fill(C_LIGHT)
            ws.cell(row=row, column=col).border = thin_border()
        c = ws.cell(row=row, column=7, value=formula)
        c.fill = fill(C_PRIMARY)
        c.font = font(size=12, bold=True, color=C_WHITE)
        c.alignment = align(h="right", v="center")
        c.border = thin_border()
        c.number_format = '"$"#,##0.00'
        row += 1


# ── Sheet 6 : Debt Tracker ────────────────────────────────────────────────────

def build_debt_tracker(wb):
    ws = wb.create_sheet("Debt Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_DANGER

    widths = {"A": 3, "B": 22, "C": 16, "D": 10, "E": 14,
              "F": 14, "G": 16, "H": 16, "I": 12, "J": 3}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    for r in range(1, 40):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 22

    # Title
    merge_and_style(ws, 2, 2, 2, 9, "DEBT PAYOFF TRACKER",
                    bg=C_DANGER, fg=C_WHITE, size=18, bold=True)
    merge_and_style(ws, 3, 2, 3, 9,
                    "Monitor your debt balances and see your payoff timeline at a glance",
                    bg=C_DARK, fg=C_WHITE, size=11)

    # Headers
    row = 5
    ws.row_dimensions[row].height = 32
    hdrs = ["Debt Name", "Original Balance", "Interest\nRate %",
            "Min Payment", "Extra Payment", "Current Balance",
            "Months to\nPayoff", "Progress %"]
    for i, h in enumerate(hdrs):
        c = header_style(ws, row, i + 2, h, bg=C_DANGER, size=11)
        c.alignment = align(h="center", v="center", wrap=True)
    row += 1

    debts = [
        ("🎓 Student Loan",   25000.00, 4.5,  285.00, 100.00, 22500.00),
        ("💳 Credit Card",     3500.00, 19.9, 70.00,  150.00,  3200.00),
        ("🚗 Car Loan",       18000.00, 3.5,  320.00,  50.00, 14800.00),
        ("💼 Personal Loan",   8000.00, 8.9,  180.00,  75.00,  6500.00),
    ]

    debt_start = row
    for name, orig, rate, min_pay, extra, curr in debts:
        ws.row_dimensions[row].height = 24
        data_cell(ws, row, 2, name, bg=C_WHITE, bold=True)

        oc = ws.cell(row=row, column=3, value=orig)
        oc.fill, oc.font = fill(C_LIGHT), font(size=11)
        oc.alignment, oc.border = align(h="right"), thin_border()
        oc.number_format = '"$"#,##0.00'

        rc = input_cell(ws, row, 4, rate / 100, num_format='0.00%')

        mc = input_cell(ws, row, 5, min_pay, num_format='"$"#,##0.00')
        ec = input_cell(ws, row, 6, extra,   num_format='"$"#,##0.00')
        cc = input_cell(ws, row, 7, curr,    num_format='"$"#,##0.00')

        # Months to payoff
        mtc = ws.cell(row=row, column=8,
                      value=f"=IF((E{row}+F{row})=0,\"N/A\",ROUND(G{row}/(E{row}+F{row}),0))")
        mtc.fill, mtc.font = fill(C_WHITE), font(size=11, bold=True)
        mtc.alignment, mtc.border = align(h="center"), thin_border()

        # Progress %
        pc = ws.cell(row=row, column=9,
                     value=f"=IF(C{row}=0,0,1-(G{row}/C{row}))")
        pc.fill, pc.font = fill(C_WHITE), font(size=11, bold=True)
        pc.alignment, pc.border = align(h="right"), thin_border()
        pc.number_format = '0.0%'
        row += 1

    debt_end = row - 1

    # Colour scale on progress
    ws.conditional_formatting.add(
        f"I{debt_start}:I{debt_end}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="E76F51",
            mid_type="num", mid_value=0.5, mid_color="E9C46A",
            end_type="num", end_value=1, end_color="2A9D8F",
        )
    )

    # Totals row
    row += 1
    ws.row_dimensions[row].height = 28
    merge_and_style(ws, row, 2, row, 2, "TOTALS", bg=C_DARK, fg=C_WHITE, size=12, bold=True)
    for col in range(3, 10):
        c = ws.cell(row=row, column=col)
        c.fill = fill(C_DARK)
        c.font = font(size=12, bold=True, color=C_WHITE)
        c.alignment = align(h="right", v="center")
        c.border = thin_border()

    # Original balance total
    ws.cell(row=row, column=3).value = f"=SUM(C{debt_start}:C{debt_end})"
    ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
    # Min payment total
    ws.cell(row=row, column=5).value = f"=SUM(E{debt_start}:E{debt_end})"
    ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
    # Extra total
    ws.cell(row=row, column=6).value = f"=SUM(F{debt_start}:F{debt_end})"
    ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
    # Current balance total
    ws.cell(row=row, column=7).value = f"=SUM(G{debt_start}:G{debt_end})"
    ws.cell(row=row, column=7).number_format = '"$"#,##0.00'

    # Tips
    row += 2
    tips = [
        ("💡 PAYOFF STRATEGY TIP:", "Pay minimums on all debts, then put every extra dollar on the highest-interest debt first (Avalanche Method)."),
        ("📅 MONTHLY REVIEW:",       "Update 'Current Balance' each month after payments to track your progress accurately."),
        ("🎯 EXTRA PAYMENTS:",        "Even $50 extra per month can save hundreds in interest and months of payments."),
    ]
    merge_and_style(ws, row, 2, row, 9, "TIPS & STRATEGIES",
                    bg=C_DARK, fg=C_WHITE, size=12, bold=True)
    row += 1
    for title, tip in tips:
        ws.row_dimensions[row].height = 36
        data_cell(ws, row, 2, title, bg=C_LIGHT, bold=True, color=C_DANGER)
        merge_and_style(ws, row, 3, row, 9, tip, bg=C_WHITE, fg="333333",
                        size=10, bold=False, h_align="left", wrap=True)
        row += 1


# ── Sheet 7 : Annual Overview ─────────────────────────────────────────────────

def build_annual_overview(wb):
    ws = wb.create_sheet("Annual Overview")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_WARNING

    # Column A = label, B-M = Jan-Dec, N = Total, O = Avg
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

    set_col_width(ws, "A", 3)
    set_col_width(ws, "B", 22)
    for i, m in enumerate(months):
        set_col_width(ws, get_column_letter(3 + i), 11)
    set_col_width(ws, get_column_letter(15), 13)  # Total
    set_col_width(ws, get_column_letter(16), 13)  # Avg

    for r in range(1, 60):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 42

    ws.freeze_panes = "C5"

    # Title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    tc = ws.cell(row=2, column=1, value="ANNUAL FINANCIAL OVERVIEW — 2026")
    tc.fill = fill(C_WARNING)
    tc.font = font(size=18, bold=True, color=C_DARK)
    tc.alignment = align(h="center", v="center")

    # Month headers
    row = 4
    ws.row_dimensions[row].height = 28
    data_cell(ws, row, 1, "", bg=C_DARK)
    header_style(ws, row, 2, "Category", bg=C_DARK, size=11)
    for i, m in enumerate(months):
        header_style(ws, row, 3 + i, m, bg=C_DARK, size=11)
    header_style(ws, row, 15, "Full Year Total", bg=C_PRIMARY, size=11)
    header_style(ws, row, 16, "Monthly Avg", bg=C_SECONDARY, size=11)

    row = 5

    def section_header(label, bg_color):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill(bg_color)
        c.font = font(size=12, bold=True, color=C_WHITE)
        c.alignment = align(h="center", v="center")
        ws.row_dimensions[row].height = 26
        row += 1

    def data_rows(items, bg_even, bg_odd):
        nonlocal row
        start_rows = []
        for i, item in enumerate(items):
            ws.row_dimensions[row].height = 22
            bg = bg_even if i % 2 == 0 else bg_odd
            data_cell(ws, row, 1, "", bg=bg)
            data_cell(ws, row, 2, item, bg=bg, bold=False)
            data_start_col = 3
            for col in range(data_start_col, data_start_col + 12):
                ic = input_cell(ws, row, col, None, num_format='"$"#,##0.00')
            # Total formula
            tc = ws.cell(row=row, column=15,
                         value=f"=SUM({get_column_letter(3)}{row}:{get_column_letter(14)}{row})")
            tc.fill, tc.font = fill(C_LIGHT), font(size=11, bold=True)
            tc.alignment, tc.border = align(h="right"), thin_border()
            tc.number_format = '"$"#,##0.00'
            # Avg formula
            ac = ws.cell(row=row, column=16,
                         value=f"=IF(O{row}=0,0,O{row}/12)")
            ac.fill, ac.font = fill(C_LIGHT), font(size=11)
            ac.alignment, ac.border = align(h="right"), thin_border()
            ac.number_format = '"$"#,##0.00'
            start_rows.append(row)
            row += 1
        return start_rows

    def total_row(label, item_rows, bg):
        nonlocal row
        ws.row_dimensions[row].height = 26
        data_cell(ws, row, 1, "", bg=bg)
        data_cell(ws, row, 2, label, bg=bg, bold=True, color=C_WHITE)
        ws.cell(row=row, column=2).fill = fill(bg)
        ws.cell(row=row, column=2).font = font(size=11, bold=True, color=C_WHITE)
        for col in range(3, 15):
            col_letter = get_column_letter(col)
            formula_parts = "+".join(f"{col_letter}{r}" for r in item_rows)
            c = ws.cell(row=row, column=col, value=f"={formula_parts}")
            c.fill = fill(bg)
            c.font = font(size=11, bold=True, color=C_WHITE)
            c.alignment = align(h="right", v="center")
            c.border = thin_border()
            c.number_format = '"$"#,##0.00'
        tc = ws.cell(row=row, column=15,
                     value=f"=SUM({get_column_letter(3)}{row}:{get_column_letter(14)}{row})")
        tc.fill, tc.font = fill(C_DARK), font(size=12, bold=True, color=C_WHITE)
        tc.alignment, tc.border = align(h="right"), thin_border()
        tc.number_format = '"$"#,##0.00'
        ac = ws.cell(row=row, column=16, value=f"=O{row}/12")
        ac.fill, ac.font = fill(C_DARK), font(size=11, bold=True, color=C_WHITE)
        ac.alignment, ac.border = align(h="right"), thin_border()
        ac.number_format = '"$"#,##0.00'
        saved = row
        row += 1
        return saved

    # INCOME section
    section_header("▶  INCOME", C_SUCCESS)
    income_items = ["Salary", "Side Hustle", "Investments", "Freelance", "Other"]
    income_rows = data_rows(income_items, C_WHITE, C_LIGHT)
    inc_total_row = total_row("TOTAL INCOME", income_rows, C_SUCCESS)

    row += 1  # spacer

    # EXPENSES section
    section_header("▶  EXPENSES", C_DANGER)
    expense_items = ["Housing", "Utilities", "Groceries", "Transportation",
                     "Insurance", "Healthcare", "Dining Out", "Entertainment",
                     "Shopping", "Subscriptions", "Savings / Emergency",
                     "Debt Payment", "Education", "Other"]
    expense_rows = data_rows(expense_items, C_WHITE, C_LIGHT)
    exp_total_row = total_row("TOTAL EXPENSES", expense_rows, C_DANGER)

    row += 1  # spacer

    # NET row
    ws.row_dimensions[row].height = 30
    data_cell(ws, row, 1, "", bg=C_PRIMARY)
    data_cell(ws, row, 2, "NET SURPLUS / DEFICIT", bg=C_PRIMARY, bold=True, color=C_WHITE)
    ws.cell(row=row, column=2).fill = fill(C_PRIMARY)
    ws.cell(row=row, column=2).font = font(size=12, bold=True, color=C_WHITE)
    for col in range(3, 15):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col,
                    value=f"={col_letter}{inc_total_row}-{col_letter}{exp_total_row}")
        c.fill = fill(C_PRIMARY)
        c.font = font(size=11, bold=True, color=C_WHITE)
        c.alignment = align(h="right", v="center")
        c.border = thin_border()
        c.number_format = '"$"#,##0.00'
    tc = ws.cell(row=row, column=15,
                 value=f"=O{inc_total_row}-O{exp_total_row}")
    tc.fill, tc.font = fill(C_DARK), font(size=13, bold=True, color=C_WHITE)
    tc.alignment, tc.border = align(h="right"), double_border_bottom()
    tc.number_format = '"$"#,##0.00'
    ac = ws.cell(row=row, column=16, value=f"=O{row}/12")
    ac.fill, ac.font = fill(C_DARK), font(size=11, bold=True, color=C_WHITE)
    ac.alignment, ac.border = align(h="right"), thin_border()
    ac.number_format = '"$"#,##0.00'


# ── Sheet 8 : Monthly Calendar ────────────────────────────────────────────────

def build_monthly_calendar(wb):
    ws = wb.create_sheet("Monthly Calendar")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_ACCENT

    # Calendar layout: cols B-H = Mon-Sun, 6 week rows
    for col in ["A", "I"]:
        set_col_width(ws, col, 2)
    for col in ["B","C","D","E","F","G","H"]:
        set_col_width(ws, col, 14)

    for r in range(1, 60):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 22
    for r in range(5, 12):
        ws.row_dimensions[r].height = 56

    # Title
    merge_and_style(ws, 2, 2, 2, 8, "JANUARY 2026 — BILL PAYMENT CALENDAR",
                    bg=C_PRIMARY, fg=C_WHITE, size=18, bold=True)
    merge_and_style(ws, 3, 2, 3, 8,
                    "Red = Bill Due  |  Green = Income Day  |  Gray = Regular Day",
                    bg=C_DARK, fg=C_WHITE, size=11)

    # Day headers
    row = 4
    ws.row_dimensions[row].height = 28
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_colors = [C_SECONDARY]*5 + [C_GRAY, C_GRAY]
    for i, (day, color) in enumerate(zip(day_names, day_colors)):
        header_style(ws, row, 2 + i, day, bg=color, size=11)

    # January 2026: 1st = Thursday → col E (index 3 = col 6 = F in B-H layout)
    # Mon=B(2), Tue=C(3), Wed=D(4), Thu=E(5), Fri=F(6), Sat=G(7), Sun=H(8)
    # Jan 1 2026 is a Thursday → position index 3 (0-based Mon-Sun)
    jan_start_dow = 3  # Thursday

    # Events
    bill_days   = {5: "Rent", 8: "Electric", 16: "Netflix", 20: "Savings Xfer", 30: "Student Loan"}
    income_days = {5: "💰 Salary", 15: "💰 Freelance"}

    row = 5
    day_num = 1
    for week in range(6):
        ws.row_dimensions[row].height = 56
        for dow in range(7):
            col = 2 + dow
            if (week == 0 and dow < jan_start_dow) or day_num > 31:
                # Empty cell
                c = ws.cell(row=row, column=col, value="")
                c.fill = fill(C_LIGHT_GRAY)
                c.border = thin_border()
            else:
                # Determine cell content and color
                day_text = str(day_num)
                cell_bg = C_LIGHT

                label_parts = [day_text]
                if day_num in bill_days:
                    label_parts.append(f"📋 {bill_days[day_num]}")
                    cell_bg = C_RED_BG
                if day_num in income_days:
                    label_parts.append(income_days[day_num])
                    cell_bg = C_GREEN_BG if day_num not in bill_days else "FFF3CD"

                cell_val = "\n".join(label_parts)
                c = ws.cell(row=row, column=col, value=cell_val)
                c.fill = fill(cell_bg)
                bold = (day_num in bill_days or day_num in income_days)
                color = C_DANGER if day_num in bill_days else (C_SUCCESS if day_num in income_days else C_DARK)
                if day_num in bill_days and day_num in income_days:
                    color = C_DARK
                c.font = font(size=10, bold=bold, color=color)
                c.alignment = align(h="center", v="top", wrap=True)
                c.border = thin_border()
                day_num += 1
        row += 1
        if day_num > 31:
            break

    # ── UPCOMING BILLS TABLE ──────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 28
    merge_and_style(ws, row, 2, row, 8, "JANUARY 2026 — UPCOMING BILLS & PAYMENTS",
                    bg=C_DARK, fg=C_WHITE, size=13, bold=True)
    row += 1
    bill_hdrs = ["Date", "Bill / Item", "Amount", "Auto-Pay?", "Status"]
    bill_col_map = [2, 3, 5, 7, 8]
    for h, col in zip(bill_hdrs, bill_col_map):
        header_style(ws, row, col, h, bg=C_PRIMARY, size=11)
    # Merge Description across cols 3-4, Amount across 5-6
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1

    bills = [
        ("Jan 5",  "🏠 Rent",           2000.00, "Yes", "Paid"),
        ("Jan 8",  "⚡ Electric Bill",     94.20, "Yes", "Paid"),
        ("Jan 15", "💰 Freelance Income", 350.00, "—",   "Received"),
        ("Jan 16", "📺 Netflix + Spotify", 26.98, "Yes", "Upcoming"),
        ("Jan 20", "🏦 Savings Transfer", 300.00, "Yes", "Upcoming"),
        ("Jan 30", "🎓 Student Loan",     350.00, "Yes", "Upcoming"),
    ]
    for date, item, amt, auto, status in bills:
        ws.row_dimensions[row].height = 24
        status_bg = C_GREEN_BG if status in ("Paid", "Received") else C_LIGHT
        status_color = C_SUCCESS if status in ("Paid", "Received") else C_ACCENT

        data_cell(ws, row, 2, date, bg=status_bg, h_align="center", bold=False)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        data_cell(ws, row, 3, item, bg=status_bg, bold=True)

        ac = ws.cell(row=row, column=5, value=amt)
        ac.fill = fill(status_bg)
        ac.font = font(size=11, bold=True)
        ac.alignment = align(h="right", v="center")
        ac.border = thin_border()
        ac.number_format = '"$"#,##0.00'
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        data_cell(ws, row, 7, auto, bg=status_bg, h_align="center")

        sc = data_cell(ws, row, 8, status, bg=status_bg, h_align="center",
                       bold=True, color=status_color)
        row += 1

    # Legend
    row += 1
    merge_and_style(ws, row, 2, row, 8, "LEGEND", bg=C_LIGHT_GRAY, fg=C_DARK, size=11, bold=True)
    row += 1
    legend = [
        (C_RED_BG,    C_DANGER,  "📋 Red Background = Bill Due Date"),
        (C_GREEN_BG,  C_SUCCESS, "💰 Green Background = Income / Payday"),
        (C_LIGHT,     C_DARK,    "⬜ Gray Background = Regular Day"),
    ]
    for bg, fg, label in legend:
        ws.row_dimensions[row].height = 24
        merge_and_style(ws, row, 2, row, 8, label, bg=bg, fg=fg,
                        size=11, bold=False, h_align="left")
        row += 1


# ── Main builder ──────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    print("Building Sheet 1: Start Here...")
    build_start_here(wb)

    print("Building Sheet 2: Category Setup...")
    build_category_setup(wb)

    print("Building Sheet 3: Monthly Tracker...")
    tracker_info = build_monthly_tracker(wb)

    print("Building Sheet 4: Dashboard...")
    build_dashboard(wb, tracker_info)

    print("Building Sheet 5: Savings Goals...")
    build_savings_goals(wb)

    print("Building Sheet 6: Debt Tracker...")
    build_debt_tracker(wb)

    print("Building Sheet 7: Annual Overview...")
    build_annual_overview(wb)

    print("Building Sheet 8: Monthly Calendar...")
    build_monthly_calendar(wb)

    # Set Start Here as the active sheet
    wb.active = wb["Start Here"]

    output_path = "Budget_Planner_2026_All_In_One.xlsx"
    wb.save(output_path)
    print(f"\n✅  Saved: {output_path}")
    return output_path


if __name__ == "__main__":
    main()

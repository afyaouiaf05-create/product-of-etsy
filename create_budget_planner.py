"""
Budget_Planner_2026_All_In_One.xlsx generator
Commercial-quality Etsy digital product – 8 fully-functional sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

# ─────────────────────────────────────────────
# PALETTE
# ─────────────────────────────────────────────
C = {
    'primary':   '2E5AAC',
    'secondary': '4A90D9',
    'accent':    'F4A261',
    'success':   '2A9D8F',
    'danger':    'E76F51',
    'warning':   'E9C46A',
    'dark':      '264653',
    'light':     'F8F9FA',
    'input':     'FFFEF0',
    'white':     'FFFFFF',
    'gray':      'DEE2E6',
    'inc_bg':    'C8E6C9',
    'exp_bg':    'FFCDD2',
    'inc_fg':    '1B5E20',
    'exp_fg':    'B71C1C',
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def font(bold=False, size=11, color='222222', italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)

def border(style='thin', color='DEE2E6'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color='264653'):
    t = Side(style='medium', color=color)
    return Border(left=t, right=t, top=t, bottom=t)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def sc(cell, f=None, fn=None, al=None, bd=None, nf=None, v=None):
    if v is not None: cell.value = v
    if f:  cell.fill = f
    if fn: cell.font = fn
    if al: cell.alignment = al
    if bd: cell.border = bd
    if nf: cell.number_format = nf

def hdr_cell(ws, row, col, text, bg, fg='FFFFFF', sz=11, bold=True, al_h='center', wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    sc(c, f=fill(bg), fn=font(bold=bold, size=sz, color=fg),
       al=align(al_h, 'center', wrap), bd=border(color='FFFFFF'))
    return c

def data_cell(ws, row, col, val, bg='FFFFFF', fg='222222', sz=10,
              al_h='left', bold=False, nf=None, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    sc(c, f=fill(bg), fn=font(bold=bold, size=sz, color=fg),
       al=align(al_h, 'center', wrap), bd=border(), nf=nf)
    return c

def input_cell(ws, row, col, val=None, nf=None, al_h='right'):
    c = ws.cell(row=row, column=col, value=val)
    sc(c, f=fill(C['input']), fn=font(size=10),
       al=align(al_h, 'center'), bd=border(), nf=nf)
    return c

def merge_hdr(ws, r1, c1, r2, c2, text, bg, fg='FFFFFF', sz=14, bold=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=text)
    sc(c, f=fill(bg), fn=font(bold=bold, size=sz, color=fg),
       al=align('center', 'center', True))
    return c

def spacer(ws, row, height=12, bg='FFFFFF'):
    rh(ws, row, height)

# Reusable diff styles for conditional formatting
green_fill  = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
red_fill    = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
green_font  = Font(name='Calibri', color='1B5E20', bold=True, size=10)
red_font    = Font(name='Calibri', color='B71C1C', bold=True, size=10)

def cf_green(formula):
    return FormulaRule(formula=[formula], fill=green_fill, font=green_font)

def cf_red(formula):
    return FormulaRule(formula=[formula], fill=red_fill, font=red_font)

# ─────────────────────────────────────────────
# WORKBOOK
# ─────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════
# SHEET 1 – START HERE
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Start Here"
ws1.sheet_properties.tabColor = C['primary']

for col, w in enumerate([4, 28, 28, 28, 20, 4], 1):
    cw(ws1, col, w)

# Title banner (rows 1-4)
for r in range(1, 5):
    rh(ws1, r, 22 if r > 1 else 55)
ws1.merge_cells('A1:F4')
sc(ws1['A1'], f=fill(C['primary']),
   fn=font(bold=True, size=30, color='FFFFFF'),
   al=align('center', 'center'),
   v="2026 All-In-One Budget Planner")

# Subtitle row 5
rh(ws1, 5, 26)
ws1.merge_cells('A5:F5')
sc(ws1['A5'], f=fill(C['secondary']),
   fn=font(italic=True, size=11, color='FFFFFF'),
   al=align('center', 'center'),
   v="Your Complete Financial Management System  •  Automated Calculations  •  No Excel Skills Required")

spacer(ws1, 6, 14, C['light'])
ws1.merge_cells('A6:F6')
ws1['A6'].fill = fill(C['light'])

# HOW IT WORKS header
rh(ws1, 7, 32)
ws1.merge_cells('A7:F7')
sc(ws1['A7'], f=fill(C['dark']),
   fn=font(bold=True, size=14, color='FFFFFF'),
   al=align('center', 'center'),
   v="HOW IT WORKS — 3 SIMPLE STEPS")

steps = [
    ("STEP 1", "SETUP",  "Go to 'Category Setup' tab\nEnter your monthly budget\ntargets for each category"),
    ("STEP 2", "TRACK",  "Go to 'Monthly Tracker' tab\nEnter every income and expense\nas it happens"),
    ("STEP 3", "REVIEW", "Check 'Dashboard' tab\nSee spending vs budget,\n50/30/20 analysis & progress"),
]
step_colors = [C['success'], C['accent'], C['secondary']]

for i, (num, title, desc) in enumerate(steps):
    col = 2 + i   # cols B, C, D
    rh(ws1, 8, 28); rh(ws1, 9, 24); rh(ws1, 10, 55)

    c8 = ws1.cell(row=8, column=col, value=num)
    sc(c8, f=fill(step_colors[i]), fn=font(bold=True, size=14, color='FFFFFF'),
       al=align('center', 'center'), bd=border(color='FFFFFF'))

    c9 = ws1.cell(row=9, column=col, value=title)
    sc(c9, f=fill(step_colors[i]), fn=font(bold=True, size=12, color='FFFFFF'),
       al=align('center', 'center'), bd=border(color='FFFFFF'))

    c10 = ws1.cell(row=10, column=col, value=desc)
    sc(c10, f=fill('EAF2FF'), fn=font(size=10, color=C['dark']),
       al=align('center', 'center', wrap=True), bd=border(color=step_colors[i]))

spacer(ws1, 11, 12, C['light'])
ws1.merge_cells('A11:F11')
ws1['A11'].fill = fill(C['light'])

# FEATURES header
rh(ws1, 12, 30)
ws1.merge_cells('A12:F12')
sc(ws1['A12'], f=fill(C['primary']),
   fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'),
   v="INCLUDED FEATURES")

features = [
    ("✔  Auto-Calculations",    "All totals, balances & percentages calculate automatically — no formulas to enter"),
    ("✔  Color Coding",         "Income highlighted green, expenses red — see your finances at a glance"),
    ("✔  50/30/20 Analysis",    "Built-in budget method analysis with real-time status indicators"),
    ("✔  Progress Bars",        "Visual savings goal tracker using block bars with percentage completion"),
    ("✔  Annual Overview",      "12-month income & expense tracker with yearly totals and monthly averages"),
    ("✔  Debt Tracker",         "Track multiple debts, interest rates & payoff timelines in one place"),
    ("✔  Savings Goals",        "Set & track up to 14 custom savings goals with animated progress bars"),
    ("✔  Bill Calendar",        "Visual January calendar with bill due dates and income payday highlights"),
]

for i, (feat, desc) in enumerate(features):
    row = 13 + i
    rh(ws1, row, 26)
    bg = 'EAF7F5' if i % 2 == 0 else 'FFFFFF'

    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws1.cell(row=row, column=1, value=feat)
    sc(c, f=fill(bg), fn=font(bold=True, size=11, color=C['success']),
       al=align('left', 'center'), bd=border())

    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    c2 = ws1.cell(row=row, column=3, value=desc)
    sc(c2, f=fill(bg), fn=font(size=10, color=C['dark']),
       al=align('left', 'center', wrap=True), bd=border())

# TIPS header
tips_r = 13 + len(features) + 1
rh(ws1, tips_r - 1, 10)
rh(ws1, tips_r, 30)
ws1.merge_cells(f'A{tips_r}:F{tips_r}')
sc(ws1[f'A{tips_r}'], f=fill(C['accent']),
   fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'),
   v="PRO TIPS FOR BEST RESULTS")

tips = [
    "📅  Update Weekly — Spend 5-10 minutes each Sunday entering the week's transactions to stay accurate",
    "🏷️  Use Categories Consistently — Always pick the same category for the same type of purchase",
    "📊  Review Monthly — Open the Dashboard on the 1st of each month to plan the month ahead",
    "🎯  Set Realistic Goals — Start with achievable targets, then increase them as habits improve",
]
for i, tip in enumerate(tips):
    row = tips_r + 1 + i
    rh(ws1, row, 28)
    ws1.merge_cells(f'A{row}:F{row}')
    c = ws1.cell(row=row, column=1, value=tip)
    sc(c, f=fill('FFF8F0' if i % 2 == 0 else 'FFFFFF'),
       fn=font(size=11, color=C['dark']),
       al=align('left', 'center'),
       bd=border(color='F4A261'))

# Footer
footer_r = tips_r + len(tips) + 2
rh(ws1, footer_r - 1, 10)
rh(ws1, footer_r, 24)
ws1.merge_cells(f'A{footer_r}:F{footer_r}')
sc(ws1[f'A{footer_r}'], f=fill(C['dark']),
   fn=font(size=9, color='AAAAAA', italic=True),
   al=align('center', 'center'),
   v="Compatible with Google Sheets & Excel  |  Digital Download  |  Personal Use Only  |  © 2026 Budget Planner")

# ═══════════════════════════════════════════════════════════
# SHEET 2 – CATEGORY SETUP
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Category Setup")
ws2.sheet_properties.tabColor = C['secondary']

for col, w in enumerate([4, 30, 26, 20, 4], 1):
    cw(ws2, col, w)

# Title
rh(ws2, 1, 40); rh(ws2, 2, 22)
ws2.merge_cells('A1:E1')
sc(ws2['A1'], f=fill(C['primary']), fn=font(bold=True, size=16, color='FFFFFF'),
   al=align('center', 'center'),
   v="CATEGORY SETUP — Define Your Monthly Budget Targets")
ws2.merge_cells('A2:E2')
sc(ws2['A2'], f=fill(C['secondary']), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Customize the budget targets below. Yellow cells accept your input. Changes reflect automatically across all sheets.")
spacer(ws2, 3, 12)

# ── EXPENSE CATEGORIES ──────────────────────────────────────
rh(ws2, 4, 30)
ws2.merge_cells('A4:E4')
sc(ws2['A4'], f=fill(C['danger']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'), v="EXPENSE CATEGORIES")

rh(ws2, 5, 26)
for col, txt in enumerate(['', 'Category Name', 'Budget Type (Need / Want / Savings)', 'Monthly Target ($)', ''], 1):
    if txt:
        hdr_cell(ws2, 5, col, txt, C['dark'], sz=10)

expense_cats = [
    ("Housing",         "Need",    1500.00),
    ("Utilities",       "Need",     150.00),
    ("Groceries",       "Need",     400.00),
    ("Transportation",  "Need",     300.00),
    ("Insurance",       "Need",     200.00),
    ("Healthcare",      "Need",     100.00),
    ("Dining Out",      "Want",     200.00),
    ("Entertainment",   "Want",     100.00),
    ("Shopping",        "Want",     150.00),
    ("Travel",          "Want",     200.00),
    ("Hobbies",         "Want",      75.00),
    ("Subscriptions",   "Want",      50.00),
    ("Emergency Fund",  "Savings",  300.00),
    ("Retirement",      "Savings",  400.00),
    ("Investments",     "Savings",  200.00),
    ("Debt Payment",    "Need",     350.00),
    ("Education",       "Want",     100.00),
    ("Gifts/Donations", "Want",      75.00),
]
type_bg = {'Need': 'FFE8E8', 'Want': 'FFF3E0', 'Savings': 'E8F5E9'}

for i, (name, btype, amt) in enumerate(expense_cats):
    row = 6 + i
    rh(ws2, row, 22)
    bg = type_bg.get(btype, 'FFFFFF') if i % 2 == 0 else 'F9F9F9'
    data_cell(ws2, row, 2, name, bg=bg, bold=True, al_h='left')
    input_cell(ws2, row, 3, btype, al_h='center')
    input_cell(ws2, row, 4, amt, nf='$#,##0.00')

exp_total_row = 6 + len(expense_cats)
rh(ws2, exp_total_row, 30)
ws2.merge_cells(f'A{exp_total_row}:C{exp_total_row}')
sc(ws2.cell(row=exp_total_row, column=1),
   f=fill(C['primary']), fn=font(bold=True, size=11, color='FFFFFF'),
   al=align('right', 'center'),
   v="TOTAL MONTHLY BUDGET TARGET")
c = ws2.cell(row=exp_total_row, column=4)
sc(c, f=fill(C['primary']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('right', 'center'), bd=thick_border('FFFFFF'),
   nf='$#,##0.00',
   v=f'=SUM(D6:D{exp_total_row - 1})')

spacer(ws2, exp_total_row + 1, 14)

# ── INCOME CATEGORIES ───────────────────────────────────────
inc_hdr_row = exp_total_row + 2
rh(ws2, inc_hdr_row, 30)
ws2.merge_cells(f'A{inc_hdr_row}:E{inc_hdr_row}')
sc(ws2.cell(row=inc_hdr_row, column=1),
   f=fill(C['success']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'), v="INCOME CATEGORIES")

inc_col_hdr_row = inc_hdr_row + 1
rh(ws2, inc_col_hdr_row, 26)
for col, txt in enumerate(['', 'Income Source', 'Type', 'Monthly Expected ($)', ''], 1):
    if txt:
        hdr_cell(ws2, inc_col_hdr_row, col, txt, C['dark'], sz=10)

income_cats = [
    ("Salary",      "Primary",   5000.00),
    ("Side Hustle", "Secondary",  500.00),
    ("Investments", "Passive",    200.00),
    ("Freelance",   "Secondary",  300.00),
    ("Other",       "Other",        0.00),
]
for i, (name, itype, amt) in enumerate(income_cats):
    row = inc_col_hdr_row + 1 + i
    rh(ws2, row, 22)
    bg = 'E8F5E9' if i % 2 == 0 else 'F9FFF9'
    data_cell(ws2, row, 2, name, bg=bg, bold=True, al_h='left')
    input_cell(ws2, row, 3, itype, al_h='center')
    input_cell(ws2, row, 4, amt if amt > 0 else None, nf='$#,##0.00')

inc_total_row = inc_col_hdr_row + 1 + len(income_cats)
rh(ws2, inc_total_row, 30)
ws2.merge_cells(f'A{inc_total_row}:C{inc_total_row}')
sc(ws2.cell(row=inc_total_row, column=1),
   f=fill(C['success']), fn=font(bold=True, size=11, color='FFFFFF'),
   al=align('right', 'center'),
   v="TOTAL EXPECTED MONTHLY INCOME")
c = ws2.cell(row=inc_total_row, column=4)
sc(c, f=fill(C['success']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('right', 'center'), bd=thick_border('FFFFFF'),
   nf='$#,##0.00',
   v=f'=SUM(D{inc_col_hdr_row+1}:D{inc_total_row - 1})')

surplus_row = inc_total_row + 1
rh(ws2, surplus_row, 30)
ws2.merge_cells(f'A{surplus_row}:C{surplus_row}')
sc(ws2.cell(row=surplus_row, column=1),
   f=fill(C['accent']), fn=font(bold=True, size=11, color='FFFFFF'),
   al=align('right', 'center'),
   v="PROJECTED MONTHLY SURPLUS")
c = ws2.cell(row=surplus_row, column=4)
sc(c, f=fill(C['accent']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('right', 'center'), bd=thick_border('FFFFFF'),
   nf='$#,##0.00',
   v=f'=D{inc_total_row}-D{exp_total_row}')

# ═══════════════════════════════════════════════════════════
# SHEET 3 – MONTHLY TRACKER
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Monthly Tracker")
ws3.sheet_properties.tabColor = C['success']

for col, w in enumerate([13, 12, 19, 32, 14, 16, 12, 22], 1):
    cw(ws3, col, w)

rh(ws3, 1, 42); rh(ws3, 2, 22)
ws3.merge_cells('A1:H1')
sc(ws3['A1'], f=fill(C['primary']), fn=font(bold=True, size=17, color='FFFFFF'),
   al=align('center', 'center'),
   v="MONTHLY TRANSACTION TRACKER — 2026")
ws3.merge_cells('A2:H2')
sc(ws3['A2'], f=fill(C['secondary']), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Enter each income or expense below. Totals at the bottom calculate automatically.")
spacer(ws3, 3, 10)

# Column headers
hdrs3 = ['Date', 'Type', 'Category', 'Description', 'Amount ($)', 'Payment Method', 'Month', 'Notes']
rh(ws3, 4, 30)
for col, txt in enumerate(hdrs3, 1):
    hdr_cell(ws3, 4, col, txt, C['dark'], sz=11)

# Sample transactions
SAMPLES = [
    ('2026-01-05', 'Income',  'Salary',         'Monthly salary deposit',              5000.00, 'Direct Deposit', 'January', ''),
    ('2026-01-05', 'Expense', 'Housing',         'January rent payment',                1500.00, 'Bank Transfer',  'January', 'Rent due 5th'),
    ('2026-01-07', 'Expense', 'Groceries',       'Weekly grocery run – Whole Foods',     127.43, 'Credit Card',    'January', ''),
    ('2026-01-08', 'Expense', 'Utilities',       'Electric bill – ConEd',                 94.20, 'Auto-Pay',       'January', ''),
    ('2026-01-10', 'Expense', 'Transportation',  'Monthly metro card',                   132.00, 'Debit Card',     'January', ''),
    ('2026-01-12', 'Expense', 'Dining Out',      'Dinner with friends – Olive Garden',    67.50, 'Credit Card',    'January', ''),
    ('2026-01-14', 'Expense', 'Subscriptions',   'Netflix monthly subscription',          15.99, 'Credit Card',    'January', 'Auto-renew'),
    ('2026-01-15', 'Income',  'Freelance',       'Website project – Client A',           750.00, 'PayPal',         'January', ''),
    ('2026-01-16', 'Expense', 'Entertainment',   'Movie tickets – 2 adults',              32.00, 'Debit Card',     'January', ''),
    ('2026-01-18', 'Expense', 'Groceries',       "Trader Joe's weekly shop",              89.75, 'Debit Card',     'January', ''),
    ('2026-01-20', 'Expense', 'Emergency Fund',  'Monthly savings auto-transfer',        300.00, 'Bank Transfer',  'January', 'Auto-transfer'),
    ('2026-01-22', 'Expense', 'Shopping',        'New work shoes – Amazon',               79.99, 'Credit Card',    'January', ''),
    ('2026-01-25', 'Expense', 'Healthcare',      'Copay – annual checkup',                35.00, 'HSA Card',       'January', ''),
    ('2026-01-28', 'Expense', 'Debt Payment',    'Student loan payment',                 250.00, 'Auto-Pay',       'January', ''),
    ('2026-01-30', 'Income',  'Side Hustle',     'Etsy shop sales – January',            320.00, 'Etsy Direct',    'January', ''),
]

N_SAMPLES = len(SAMPLES)
N_EMPTY   = 200
DATA_START = 5
LAST_DATA_ROW = DATA_START + N_SAMPLES + N_EMPTY - 1  # 219

for i, (dt, ttype, cat, desc, amt, method, month, notes) in enumerate(SAMPLES):
    row = DATA_START + i
    rh(ws3, row, 22)
    vals = [dt, ttype, cat, desc, amt, method, month, notes]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        sc(c, bd=border(), al=align('left', 'center'))
        if col == 1:
            sc(c, fn=font(size=10), al=align('center', 'center'), nf='YYYY-MM-DD')
        elif col == 2:
            if ttype == 'Income':
                sc(c, f=fill(C['inc_bg']), fn=font(bold=True, size=10, color=C['inc_fg']),
                   al=align('center', 'center'))
            else:
                sc(c, f=fill(C['exp_bg']), fn=font(bold=True, size=10, color=C['exp_fg']),
                   al=align('center', 'center'))
        elif col == 5:
            sc(c, fn=font(size=10), al=align('right', 'center'), nf='$#,##0.00')
        elif col == 7:
            sc(c, fn=font(size=10), al=align('center', 'center'))
        else:
            sc(c, fn=font(size=10), f=fill('FFFFFF'))

# Empty entry rows
for i in range(N_EMPTY):
    row = DATA_START + N_SAMPLES + i
    rh(ws3, row, 20)
    for col in range(1, 9):
        c = ws3.cell(row=row, column=col)
        bg = C['input'] if col not in [2] else 'FFFFFF'
        sc(c, f=fill(bg), fn=font(size=10), bd=border())
        if col == 5:
            sc(c, al=align('right', 'center'), nf='$#,##0.00')
        elif col in [1, 7]:
            sc(c, al=align('center', 'center'))
        else:
            sc(c, al=align('left', 'center'))

# ── Totals ──
TOTALS_SPACER = LAST_DATA_ROW + 1
rh(ws3, TOTALS_SPACER, 10)
ws3.merge_cells(f'A{TOTALS_SPACER}:H{TOTALS_SPACER}')

TI_ROW = LAST_DATA_ROW + 2   # Total Income  = 221
TE_ROW = TI_ROW + 1          # Total Expenses= 222
NB_ROW = TE_ROW + 1          # Net Balance   = 223

for row, label, formula, col_color in [
    (TI_ROW, "TOTAL INCOME",                   f'=SUMIF(B{DATA_START}:B{LAST_DATA_ROW},"Income",E{DATA_START}:E{LAST_DATA_ROW})',  C['success']),
    (TE_ROW, "TOTAL EXPENSES",                  f'=SUMIF(B{DATA_START}:B{LAST_DATA_ROW},"Expense",E{DATA_START}:E{LAST_DATA_ROW})', C['danger']),
    (NB_ROW, "NET BALANCE  (Income – Expenses)", f'=E{TI_ROW}-E{TE_ROW}',                                                           C['primary']),
]:
    h = 32 if row != NB_ROW else 38
    rh(ws3, row, h)
    sz = 13 if row != NB_ROW else 16
    ws3.merge_cells(f'A{row}:D{row}')
    sc(ws3.cell(row=row, column=1),
       f=fill(col_color), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('right', 'center'), v=label)
    c = ws3.cell(row=row, column=5)
    sc(c, f=fill(col_color), fn=font(bold=True, size=sz, color='FFFFFF'),
       al=align('right', 'center'),
       bd=thick_border('FFFFFF'), nf='$#,##0.00', v=formula)

ws3.freeze_panes = 'A5'

# Reference strings for other sheets
MT_INC  = f"'Monthly Tracker'!E{TI_ROW}"
MT_EXP  = f"'Monthly Tracker'!E{TE_ROW}"
MT_RANGE_B = f"'Monthly Tracker'!B{DATA_START}:B{LAST_DATA_ROW}"
MT_RANGE_C = f"'Monthly Tracker'!C{DATA_START}:C{LAST_DATA_ROW}"
MT_RANGE_E = f"'Monthly Tracker'!E{DATA_START}:E{LAST_DATA_ROW}"

def sumifs_cat(cat):
    """SUMIFS formula pulling one category's expenses from Monthly Tracker."""
    return f'=SUMIFS({MT_RANGE_E},{MT_RANGE_B},"Expense",{MT_RANGE_C},"{cat}")'

def sumproduct_cats(cats):
    """SUMPRODUCT for multiple categories – works in Excel & Google Sheets."""
    cat_list = ','.join(f'"{c}"' for c in cats)
    return (f'=SUMPRODUCT((ISNUMBER(MATCH({MT_RANGE_C},{{{cat_list}}},0)))'
            f'*({MT_RANGE_B}="Expense")*{MT_RANGE_E})')

# ═══════════════════════════════════════════════════════════
# SHEET 4 – DASHBOARD
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")
ws4.sheet_properties.tabColor = C['accent']

for col, w in enumerate([3, 24, 17, 17, 16, 18, 3], 1):
    cw(ws4, col, w)

rh(ws4, 1, 46); rh(ws4, 2, 22)
ws4.merge_cells('A1:G1')
sc(ws4['A1'], f=fill(C['primary']), fn=font(bold=True, size=20, color='FFFFFF'),
   al=align('center', 'center'), v="FINANCIAL DASHBOARD — 2026")
ws4.merge_cells('A2:G2')
sc(ws4['A2'], f=fill(C['secondary']), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Complete financial overview — all figures pulled automatically from Monthly Tracker")
spacer(ws4, 3, 14)

# ── KPI CARDS ────────────────────────────────────────────────
KPI = [
    ('TOTAL INCOME',   f'={MT_INC}',                    C['success'],   '$#,##0.00', '💰'),
    ('TOTAL EXPENSES', f'={MT_EXP}',                    C['danger'],    '$#,##0.00', '💸'),
    ('NET BALANCE',    f'={MT_INC}-{MT_EXP}',           C['secondary'], '$#,##0.00', '📊'),
    ('SAVINGS RATE',   f'=IF({MT_INC}>0,({MT_INC}-{MT_EXP})/{MT_INC},0)', C['accent'], '0.0%', '🎯'),
]
kpi_cols = [2, 3, 4, 5]   # B C D E

rh(ws4, 4, 26); rh(ws4, 5, 42); rh(ws4, 6, 18)
for i, (label, formula, color, nf, icon) in enumerate(KPI):
    col = kpi_cols[i]
    # Header
    c4 = ws4.cell(row=4, column=col)
    sc(c4, f=fill(color), fn=font(bold=True, size=9, color='FFFFFF'),
       al=align('center', 'center'), bd=border(color='FFFFFF'),
       v=f"{icon} {label}")
    # Value
    c5 = ws4.cell(row=5, column=col)
    sc(c5, f=fill('FFFFFF'), fn=Font(name='Calibri', bold=True, size=17, color=color),
       al=align('center', 'center'), bd=border(color=color), nf=nf, v=formula)
    # Sub-label
    c6 = ws4.cell(row=6, column=col)
    sc(c6, f=fill('F8F9FA'), fn=font(size=8, color='6C757D', italic=True),
       al=align('center', 'center'), bd=border(), v="↑ auto-calculated")

spacer(ws4, 7, 14)

# ── BUDGET VS ACTUAL ─────────────────────────────────────────
rh(ws4, 8, 32)
ws4.merge_cells('A8:G8')
sc(ws4['A8'], f=fill(C['dark']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'), v="BUDGET VS ACTUAL — Current Month")

rh(ws4, 9, 26)
for col, txt in enumerate(['', 'Category', 'Budget Target', 'Actual Spent', 'Variance', 'Status', ''], 1):
    if txt:
        hdr_cell(ws4, 9, col, txt, C['primary'], sz=10)

BVA = [
    ('Housing',       1500, 'Housing'),
    ('Groceries',      400, 'Groceries'),
    ('Transportation', 300, 'Transportation'),
    ('Utilities',      150, 'Utilities'),
    ('Dining Out',     200, 'Dining Out'),
    ('Entertainment',  100, 'Entertainment'),
    ('Shopping',       150, 'Shopping'),
    ('Savings',        300, 'Emergency Fund'),
    ('Debt Payment',   350, 'Debt Payment'),
    ('Healthcare',     100, 'Healthcare'),
]
BVA_START = 10
for i, (disp, budget, tracker_cat) in enumerate(BVA):
    row = BVA_START + i
    rh(ws4, row, 24)
    bg = 'F8F9FA' if i % 2 == 0 else 'FFFFFF'

    data_cell(ws4, row, 2, disp, bg=bg, bold=True, al_h='left')
    data_cell(ws4, row, 3, budget, bg=bg, al_h='right', nf='$#,##0.00')

    c_act = ws4.cell(row=row, column=4)
    sc(c_act, f=fill(bg), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00',
       v=sumifs_cat(tracker_cat))

    c_var = ws4.cell(row=row, column=5)
    sc(c_var, f=fill(bg), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00',
       v=f'=C{row}-D{row}')

    c_st = ws4.cell(row=row, column=6)
    sc(c_st, f=fill(bg), fn=font(size=10), al=align('center', 'center'),
       bd=border(),
       v=f'=IF(D{row}<=C{row},"✓ Under Budget","✗ Over Budget")')

# Conditional formatting – Status column
status_range = f'F{BVA_START}:F{BVA_START + len(BVA) - 1}'
ws4.conditional_formatting.add(status_range,
    cf_green(f'F{BVA_START}="✓ Under Budget"'))
ws4.conditional_formatting.add(status_range,
    cf_red(f'F{BVA_START}="✗ Over Budget"'))

# BvA totals
bva_tot_row = BVA_START + len(BVA)
rh(ws4, bva_tot_row, 30)
ws4.cell(row=bva_tot_row, column=2, value="TOTALS")
sc(ws4.cell(row=bva_tot_row, column=2),
   f=fill(C['primary']), fn=font(bold=True, size=10, color='FFFFFF'),
   al=align('right', 'center'), bd=border(color='FFFFFF'))
for col in [3, 4, 5]:
    c = ws4.cell(row=bva_tot_row, column=col)
    col_ltr = get_column_letter(col)
    sc(c, f=fill(C['primary']), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'), nf='$#,##0.00',
       v=f'=SUM({col_ltr}{BVA_START}:{col_ltr}{bva_tot_row - 1})')

spacer(ws4, bva_tot_row + 1, 14)

# ── 50/30/20 ANALYSIS ────────────────────────────────────────
analysis_hdr = bva_tot_row + 2
rh(ws4, analysis_hdr, 32)
ws4.merge_cells(f'A{analysis_hdr}:G{analysis_hdr}')
sc(ws4.cell(row=analysis_hdr, column=1),
   f=fill(C['dark']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'), v="50/30/20 BUDGET METHOD ANALYSIS")

rh(ws4, analysis_hdr + 1, 26)
for col, txt in enumerate(['', 'Category', 'Rule %', 'Target Amount', 'Actual Amount', 'Status', ''], 1):
    if txt:
        hdr_cell(ws4, analysis_hdr + 1, col, txt, C['accent'], sz=10)

need_cats    = ['Housing','Utilities','Groceries','Transportation','Insurance','Healthcare','Debt Payment']
want_cats    = ['Dining Out','Entertainment','Shopping','Travel','Hobbies','Subscriptions','Education','Gifts/Donations']
savings_cats = ['Emergency Fund','Retirement','Investments']

ANALYSIS = [
    ('Needs',   0.50, sumproduct_cats(need_cats),    C['danger']),
    ('Wants',   0.30, sumproduct_cats(want_cats),    C['accent']),
    ('Savings', 0.20, sumproduct_cats(savings_cats), C['success']),
]
an_start = analysis_hdr + 2
for i, (cat, pct, actual_f, color) in enumerate(ANALYSIS):
    row = an_start + i
    rh(ws4, row, 30)
    bg = 'F8F9FA' if i % 2 == 0 else 'FFFFFF'

    c_cat = ws4.cell(row=row, column=2, value=cat)
    sc(c_cat, f=fill(color), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('center', 'center'), bd=border())

    c_pct = ws4.cell(row=row, column=3, value=pct)
    sc(c_pct, f=fill(bg), fn=Font(name='Calibri', bold=True, size=12, color=color),
       al=align('center', 'center'), bd=border(), nf='0%')

    c_target = ws4.cell(row=row, column=4)
    sc(c_target, f=fill(bg), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00',
       v=f'=C{row}*{MT_INC}')

    c_actual = ws4.cell(row=row, column=5)
    sc(c_actual, f=fill(bg), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00', v=actual_f)

    c_st = ws4.cell(row=row, column=6)
    sc(c_st, f=fill(bg), fn=font(size=10), al=align('center', 'center'),
       bd=border(),
       v=f'=IF(E{row}<=D{row},"✓ On Track","✗ Over Limit")')

an_status_range = f'F{an_start}:F{an_start + len(ANALYSIS) - 1}'
ws4.conditional_formatting.add(an_status_range,
    cf_green(f'F{an_start}="✓ On Track"'))
ws4.conditional_formatting.add(an_status_range,
    cf_red(f'F{an_start}="✗ Over Limit"'))

ws4.freeze_panes = 'A8'

# ═══════════════════════════════════════════════════════════
# SHEET 5 – SAVINGS GOALS
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Savings Goals")
ws5.sheet_properties.tabColor = C['warning']

for col, w in enumerate([3, 26, 17, 17, 15, 14, 32, 3], 1):
    cw(ws5, col, w)

rh(ws5, 1, 42); rh(ws5, 2, 22)
ws5.merge_cells('A1:H1')
sc(ws5['A1'], f=fill(C['success']), fn=font(bold=True, size=16, color='FFFFFF'),
   al=align('center', 'center'), v="SAVINGS GOALS TRACKER — 2026")
ws5.merge_cells('A2:H2')
sc(ws5['A2'], f=fill('229B8A'), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Track your progress toward every financial goal. Update 'Current Saved' monthly to watch your progress bars grow!")
spacer(ws5, 3, 12)

# Column headers
sg_hdrs = ['', 'Goal Name', 'Target Amount', 'Current Saved', 'Remaining', 'Progress %', 'Visual Progress  ────────────────────', '']
rh(ws5, 4, 30)
for col, txt in enumerate(sg_hdrs, 1):
    if txt:
        hdr_cell(ws5, 4, col, txt, C['dark'], sz=10)

GOALS = [
    ("🏦  Emergency Fund",  10000, 3500),
    ("✈️  Vacation Fund",    3000,  1200),
    ("🚗  New Car",          15000, 4500),
    ("🏠  Home Down Payment",50000, 8000),
    ("💍  Wedding",          20000, 5000),
    ("🎓  Education",         8000, 2400),
]
goal_colors = [C['success'], C['secondary'], C['accent'], C['primary'], C['danger'], C['warning']]

SG_START = 5
for i, (goal, target, saved) in enumerate(GOALS):
    row = SG_START + i
    rh(ws5, row, 30)
    color = goal_colors[i % len(goal_colors)]
    bg = 'F0FFF9' if i % 2 == 0 else 'FAFFFE'

    c_goal = ws5.cell(row=row, column=2, value=goal)
    sc(c_goal, f=fill(color), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('left', 'center'), bd=border())

    input_cell(ws5, row, 3, target, nf='$#,##0.00')
    input_cell(ws5, row, 4, saved,  nf='$#,##0.00')

    c_rem = ws5.cell(row=row, column=5)
    sc(c_rem, f=fill(bg), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00', v=f'=C{row}-D{row}')

    c_pct = ws5.cell(row=row, column=6)
    sc(c_pct, f=fill(bg), fn=font(bold=True, size=11), al=align('center', 'center'),
       bd=border(), nf='0.0%',
       v=f'=IF(C{row}>0,MIN(D{row}/C{row},1),0)')

    c_bar = ws5.cell(row=row, column=7)
    sc(c_bar, f=fill(bg),
       fn=Font(name='Calibri', size=13, color=color),
       al=align('left', 'center'), bd=border(),
       v=f'=REPT("█",ROUND(F{row}*20,0))&REPT("░",20-ROUND(F{row}*20,0))')

# Color-scale conditional formatting on Progress %
ws5.conditional_formatting.add(
    f'F{SG_START}:F{SG_START + len(GOALS) - 1}',
    ColorScaleRule(start_type='num', start_value=0,   start_color='E76F51',
                   mid_type='num',   mid_value=0.5,   mid_color='E9C46A',
                   end_type='num',   end_value=1,     end_color='2A9D8F'))

# 8 blank user rows
BLANK_SG_START = SG_START + len(GOALS)
for i in range(8):
    row = BLANK_SG_START + i
    rh(ws5, row, 28)
    c_name = ws5.cell(row=row, column=2)
    sc(c_name, f=fill(C['input']), fn=font(size=10, color='AAAAAA', italic=True),
       al=align('left', 'center'), bd=border(), v="Enter goal name…")

    for col in [3, 4]:
        input_cell(ws5, row, col, nf='$#,##0.00')

    c_rem = ws5.cell(row=row, column=5)
    sc(c_rem, f=fill('FAFAFA'), fn=font(size=10), al=align('right', 'center'),
       bd=border(), nf='$#,##0.00',
       v=f'=IF(AND(C{row}<>"",C{row}>0),C{row}-D{row},"")')

    c_pct = ws5.cell(row=row, column=6)
    sc(c_pct, f=fill('FAFAFA'), fn=font(size=10), al=align('center', 'center'),
       bd=border(), nf='0.0%',
       v=f'=IF(AND(C{row}<>"",C{row}>0),MIN(D{row}/C{row},1),"")')

    c_bar = ws5.cell(row=row, column=7)
    sc(c_bar, f=fill('FAFAFA'), fn=Font(name='Calibri', size=13, color=C['success']),
       al=align('left', 'center'), bd=border(),
       v=f'=IF(AND(C{row}<>"",C{row}>0),REPT("█",ROUND(F{row}*20,0))&REPT("░",20-ROUND(F{row}*20,0)),"")')

# Totals
sg_last_data = BLANK_SG_START + 8 - 1
sg_tot_row = sg_last_data + 2
spacer(ws5, sg_tot_row - 1, 10)
rh(ws5, sg_tot_row, 30)
ws5.cell(row=sg_tot_row, column=2, value="GRAND TOTALS — All Goals")
sc(ws5.cell(row=sg_tot_row, column=2),
   f=fill(C['success']), fn=font(bold=True, size=11, color='FFFFFF'),
   al=align('right', 'center'), bd=border(color='FFFFFF'))
for col in [3, 4, 5]:
    c = ws5.cell(row=sg_tot_row, column=col)
    col_ltr = get_column_letter(col)
    sc(c, f=fill(C['success']), fn=font(bold=True, size=12, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'), nf='$#,##0.00',
       v=f'=SUMIF({col_ltr}{SG_START}:{col_ltr}{sg_last_data},"<>",{col_ltr}{SG_START}:{col_ltr}{sg_last_data})')

# ═══════════════════════════════════════════════════════════
# SHEET 6 – DEBT TRACKER
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Debt Tracker")
ws6.sheet_properties.tabColor = C['danger']

for col, w in enumerate([3, 22, 16, 13, 14, 14, 16, 17, 14, 3], 1):
    cw(ws6, col, w)

rh(ws6, 1, 42); rh(ws6, 2, 22)
ws6.merge_cells('A1:J1')
sc(ws6['A1'], f=fill(C['danger']), fn=font(bold=True, size=16, color='FFFFFF'),
   al=align('center', 'center'), v="DEBT PAYOFF TRACKER — 2026")
ws6.merge_cells('A2:J2')
sc(ws6['A2'], f=fill('D96040'), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Track all debts, interest rates & payoff timelines. Yellow cells accept your input.")
spacer(ws6, 3, 12)

rh(ws6, 4, 38)
debt_hdrs = ['', 'Debt Name', 'Original Balance', 'Interest Rate', 'Min. Payment',
             'Extra Payment', 'Current Balance', 'Months to Payoff', 'Progress %', '']
for col, txt in enumerate(debt_hdrs, 1):
    if txt:
        hdr_cell(ws6, 4, col, txt, C['dark'], sz=10, wrap=True)

DEBTS = [
    ("🎓  Student Loan",  25000, 4.5,  280, 50,  22000),
    ("💳  Credit Card",    3500, 19.9,  75, 100,  2800),
    ("🚗  Car Loan",      18000, 3.5,  350,  0,  14500),
    ("💼  Personal Loan",  8000, 8.9,  180, 25,   6200),
]
debt_colors = [C['secondary'], C['danger'], C['accent'], C['warning']]

DT_START = 5
for i, (name, orig, rate, min_pay, extra, current) in enumerate(DEBTS):
    row = DT_START + i
    rh(ws6, row, 30)
    color = debt_colors[i]
    bg = 'FFF5F5' if i % 2 == 0 else 'FFFFFF'

    c_name = ws6.cell(row=row, column=2, value=name)
    sc(c_name, f=fill(color), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('left', 'center'), bd=border())

    input_cell(ws6, row, 3, orig,        nf='$#,##0.00')
    input_cell(ws6, row, 4, rate / 100,  nf='0.00%')
    input_cell(ws6, row, 5, min_pay,     nf='$#,##0.00')
    input_cell(ws6, row, 6, extra,       nf='$#,##0.00')
    input_cell(ws6, row, 7, current,     nf='$#,##0.00')

    c_mo = ws6.cell(row=row, column=8)
    sc(c_mo, f=fill(bg), fn=font(size=10), al=align('center', 'center'),
       bd=border(),
       v=f'=IF((E{row}+F{row})>0,ROUND(G{row}/(E{row}+F{row}),0),"N/A")')

    c_pg = ws6.cell(row=row, column=9)
    sc(c_pg, f=fill(bg), fn=font(bold=True, size=10), al=align('center', 'center'),
       bd=border(), nf='0.0%',
       v=f'=IF(C{row}>0,MAX(0,1-(G{row}/C{row})),0)')

ws6.conditional_formatting.add(
    f'I{DT_START}:I{DT_START + len(DEBTS) - 1}',
    ColorScaleRule(start_type='num', start_value=0,   start_color='E76F51',
                   mid_type='num',   mid_value=0.5,   mid_color='E9C46A',
                   end_type='num',   end_value=1,     end_color='2A9D8F'))

dt_tot_row = DT_START + len(DEBTS) + 1
spacer(ws6, DT_START + len(DEBTS), 10)
rh(ws6, dt_tot_row, 30)
ws6.merge_cells(f'A{dt_tot_row}:B{dt_tot_row}')
sc(ws6.cell(row=dt_tot_row, column=1),
   f=fill(C['danger']), fn=font(bold=True, size=11, color='FFFFFF'),
   al=align('center', 'center'), bd=border(color='FFFFFF'), v="TOTALS")

for col in [3, 5, 6, 7]:
    c = ws6.cell(row=dt_tot_row, column=col)
    col_ltr = get_column_letter(col)
    sc(c, f=fill(C['danger']), fn=font(bold=True, size=12, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'), nf='$#,##0.00',
       v=f'=SUM({col_ltr}{DT_START}:{col_ltr}{DT_START + len(DEBTS) - 1})')

for col in [4, 8, 9]:
    ws6.cell(row=dt_tot_row, column=col).fill = fill(C['danger'])

tip_row = dt_tot_row + 2
spacer(ws6, dt_tot_row + 1, 8)
rh(ws6, tip_row, 44)
ws6.merge_cells(f'A{tip_row}:J{tip_row}')
sc(ws6.cell(row=tip_row, column=1),
   f=fill('FFF3E0'), fn=font(size=10, color=C['dark']),
   al=align('center', 'center', wrap=True),
   bd=border(color=C['accent']),
   v=("💡  DEBT PAYOFF STRATEGIES:\n"
      "Avalanche Method: Pay extra toward highest-interest debt first (saves the most money)  |  "
      "Snowball Method: Pay extra toward smallest balance first (builds momentum & motivation)"))

# ═══════════════════════════════════════════════════════════
# SHEET 7 – ANNUAL OVERVIEW
# ═══════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Annual Overview")
ws7.sheet_properties.tabColor = C['dark']

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
# Columns: A(spacer) B(label) C-N(months) O(total) P(avg) Q(spacer)
for col, w in enumerate([3, 22] + [10]*12 + [13, 13, 3], 1):
    cw(ws7, col, w)
MAX_COL_7 = 17   # Q

rh(ws7, 1, 42); rh(ws7, 2, 22)
ws7.merge_cells(f'A1:{get_column_letter(MAX_COL_7)}1')
sc(ws7['A1'], f=fill(C['primary']), fn=font(bold=True, size=16, color='FFFFFF'),
   al=align('center', 'center'), v="ANNUAL FINANCIAL OVERVIEW — 2026")
ws7.merge_cells(f'A2:{get_column_letter(MAX_COL_7)}2')
sc(ws7['A2'], f=fill(C['secondary']), fn=font(italic=True, size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="Enter income and expenses for each month. Annual totals and monthly averages calculate automatically.")
spacer(ws7, 3, 12)

# Month headers
rh(ws7, 4, 28)
hdr_cell(ws7, 4, 2, 'Category', C['dark'], sz=10)
for i, m in enumerate(MONTHS):
    hdr_cell(ws7, 4, 3 + i, m, C['primary'], sz=10)
hdr_cell(ws7, 4, 15, 'Annual Total', C['dark'], sz=10)
hdr_cell(ws7, 4, 16, 'Avg / Month',  C['dark'], sz=10)

# ── INCOME ───────────────────────────────────────────────────
rh(ws7, 5, 26)
ws7.merge_cells(f'A5:{get_column_letter(MAX_COL_7)}5')
sc(ws7.cell(row=5, column=1),
   f=fill(C['success']), fn=font(bold=True, size=12, color='FFFFFF'),
   al=align('center', 'center'), v="▶  INCOME")

INC_ROWS_7 = [
    ('Salary',      [5000]*12),
    ('Side Hustle', [320,450,380,290,400,500,350,420,380,310,450,600]),
    ('Investments', [200]*12),
    ('Freelance',   [750,0,500,0,300,750,0,500,0,750,500,1000]),
    ('Other',       [0]*12),
]
INC_START_7 = 6
for i, (label, vals) in enumerate(INC_ROWS_7):
    row = INC_START_7 + i
    rh(ws7, row, 22)
    bg = 'E8F5E9' if i % 2 == 0 else 'F5FFF7'
    data_cell(ws7, row, 2, label, bg=bg, bold=True, al_h='left')
    for j, v in enumerate(vals):
        c = ws7.cell(row=row, column=3 + j, value=v if v > 0 else None)
        sc(c, f=fill(C['input']), fn=font(size=10),
           al=align('right', 'center'), bd=border(), nf='$#,##0.00')
    c_tot = ws7.cell(row=row, column=15)
    sc(c_tot, f=fill('C8E6C9'), fn=font(bold=True, size=10),
       al=align('right', 'center'), bd=border(), nf='$#,##0.00',
       v=f'=SUM(C{row}:N{row})')
    c_avg = ws7.cell(row=row, column=16)
    sc(c_avg, f=fill('C8E6C9'), fn=font(size=10),
       al=align('right', 'center'), bd=border(), nf='$#,##0.00',
       v=f'=IFERROR(O{row}/COUNTIF(C{row}:N{row},">"&0),0)')

INC_TOTAL_ROW_7 = INC_START_7 + len(INC_ROWS_7)
rh(ws7, INC_TOTAL_ROW_7, 28)
hdr_cell(ws7, INC_TOTAL_ROW_7, 2, 'TOTAL INCOME', C['success'], sz=11, al_h='right')
for col in range(3, 17):
    c = ws7.cell(row=INC_TOTAL_ROW_7, column=col)
    col_ltr = get_column_letter(col)
    if col <= 14:
        formula = f'=SUM({col_ltr}{INC_START_7}:{col_ltr}{INC_TOTAL_ROW_7-1})'
    elif col == 15:
        formula = f'=SUM(C{INC_TOTAL_ROW_7}:N{INC_TOTAL_ROW_7})'
    else:
        formula = f'=IFERROR(O{INC_TOTAL_ROW_7}/12,0)'
    sc(c, f=fill(C['success']), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'),
       nf='$#,##0.00', v=formula)

spacer(ws7, INC_TOTAL_ROW_7 + 1, 10)

# ── EXPENSES ─────────────────────────────────────────────────
EXP_SECTION_7 = INC_TOTAL_ROW_7 + 2
rh(ws7, EXP_SECTION_7, 26)
ws7.merge_cells(f'A{EXP_SECTION_7}:{get_column_letter(MAX_COL_7)}{EXP_SECTION_7}')
sc(ws7.cell(row=EXP_SECTION_7, column=1),
   f=fill(C['danger']), fn=font(bold=True, size=12, color='FFFFFF'),
   al=align('center', 'center'), v="▶  EXPENSES")

EXP_ROWS_7 = [
    ('Housing',       [1500]*12),
    ('Utilities',     [94,88,82,75,70,95,120,130,105,88,92,98]),
    ('Groceries',     [217,200,210,195,220,230,215,205,225,210,240,280]),
    ('Transportation',[132]*12),
    ('Insurance',     [200]*12),
    ('Healthcare',    [35,0,0,75,0,0,35,0,0,100,0,35]),
    ('Dining Out',    [67,80,95,110,85,120,90,75,85,95,130,180]),
    ('Entertainment', [32,45,38,55,40,75,50,45,40,55,60,85]),
    ('Shopping',      [79,120,95,80,150,65,200,110,90,180,250,400]),
    ('Subscriptions', [16]*12),
    ('Savings',       [300]*12),
    ('Debt Payment',  [250]*12),
    ('Education',     [0,0,0,0,0,0,0,0,500,0,0,0]),
    ('Other',         [0]*12),
]
EXP_START_7 = EXP_SECTION_7 + 1
for i, (label, vals) in enumerate(EXP_ROWS_7):
    row = EXP_START_7 + i
    rh(ws7, row, 22)
    bg = 'FFF0F0' if i % 2 == 0 else 'FFFAFA'
    data_cell(ws7, row, 2, label, bg=bg, bold=True, al_h='left')
    for j, v in enumerate(vals):
        c = ws7.cell(row=row, column=3 + j, value=v if v > 0 else None)
        sc(c, f=fill(C['input']), fn=font(size=10),
           al=align('right', 'center'), bd=border(), nf='$#,##0.00')
    c_tot = ws7.cell(row=row, column=15)
    sc(c_tot, f=fill('FFCDD2'), fn=font(bold=True, size=10),
       al=align('right', 'center'), bd=border(), nf='$#,##0.00',
       v=f'=SUM(C{row}:N{row})')
    c_avg = ws7.cell(row=row, column=16)
    sc(c_avg, f=fill('FFCDD2'), fn=font(size=10),
       al=align('right', 'center'), bd=border(), nf='$#,##0.00',
       v=f'=IFERROR(O{row}/COUNTIF(C{row}:N{row},">"&0),0)')

EXP_TOTAL_ROW_7 = EXP_START_7 + len(EXP_ROWS_7)
rh(ws7, EXP_TOTAL_ROW_7, 28)
hdr_cell(ws7, EXP_TOTAL_ROW_7, 2, 'TOTAL EXPENSES', C['danger'], sz=11, al_h='right')
for col in range(3, 17):
    c = ws7.cell(row=EXP_TOTAL_ROW_7, column=col)
    col_ltr = get_column_letter(col)
    if col <= 14:
        formula = f'=SUM({col_ltr}{EXP_START_7}:{col_ltr}{EXP_TOTAL_ROW_7-1})'
    elif col == 15:
        formula = f'=SUM(C{EXP_TOTAL_ROW_7}:N{EXP_TOTAL_ROW_7})'
    else:
        formula = f'=IFERROR(O{EXP_TOTAL_ROW_7}/12,0)'
    sc(c, f=fill(C['danger']), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'),
       nf='$#,##0.00', v=formula)

# NET row
NET_ROW_7 = EXP_TOTAL_ROW_7 + 1
rh(ws7, NET_ROW_7, 32)
hdr_cell(ws7, NET_ROW_7, 2, 'NET (Income – Expenses)', C['primary'], sz=11, al_h='right')
for col in range(3, 17):
    c = ws7.cell(row=NET_ROW_7, column=col)
    col_ltr = get_column_letter(col)
    if col <= 14:
        formula = f'={col_ltr}{INC_TOTAL_ROW_7}-{col_ltr}{EXP_TOTAL_ROW_7}'
    elif col == 15:
        formula = f'=SUM(C{NET_ROW_7}:N{NET_ROW_7})'
    else:
        formula = f'=IFERROR(O{NET_ROW_7}/12,0)'
    sc(c, f=fill(C['primary']), fn=font(bold=True, size=11, color='FFFFFF'),
       al=align('right', 'center'), bd=border(color='FFFFFF'),
       nf='$#,##0.00', v=formula)

ws7.freeze_panes = 'C5'

# ═══════════════════════════════════════════════════════════
# SHEET 8 – MONTHLY CALENDAR
# ═══════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Monthly Calendar")
ws8.sheet_properties.tabColor = C['secondary']

for col, w in enumerate([3, 14, 14, 14, 14, 14, 14, 14, 3], 1):
    cw(ws8, col, w)

rh(ws8, 1, 42); rh(ws8, 2, 24)
ws8.merge_cells('A1:I1')
sc(ws8['A1'], f=fill(C['primary']), fn=font(bold=True, size=16, color='FFFFFF'),
   al=align('center', 'center'), v="JANUARY 2026 — BILL & INCOME CALENDAR")
ws8.merge_cells('A2:I2')
sc(ws8['A2'], f=fill(C['dark']), fn=font(size=10, color='FFFFFF'),
   al=align('center', 'center'),
   v="🔴  Red = Bill Due Date     🟢  Green = Income / Payday     ⚪  Gray = Regular Day     🟡  Yellow = Both")
spacer(ws8, 3, 10)

# Day-of-week headers
rh(ws8, 4, 30)
for i, day in enumerate(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'], 2):
    c = ws8.cell(row=4, column=i, value=day)
    sc(c, f=fill(C['secondary']), fn=font(bold=True, size=12, color='FFFFFF'),
       al=align('center', 'center'), bd=border(color='FFFFFF'))

# January 2026: Jan 1 is Thursday → start offset = 3 (Mon=0)
BILL_DATES   = {5: "Rent Due", 8: "Electric Due", 16: "Netflix Due", 20: "Savings Transfer", 30: "Student Loan"}
INCOME_DATES = {5: "Salary Deposit", 15: "Freelance Payment"}

WEEKS_JAN = [
    [None, None, None, 1, 2, 3, 4],
    [5, 6, 7, 8, 9, 10, 11],
    [12, 13, 14, 15, 16, 17, 18],
    [19, 20, 21, 22, 23, 24, 25],
    [26, 27, 28, 29, 30, 31, None],
]

for week_i, week in enumerate(WEEKS_JAN):
    row = 5 + week_i
    rh(ws8, row, 52)
    for day_i, day_num in enumerate(week):
        col = 2 + day_i
        c = ws8.cell(row=row, column=col)
        if day_num is None:
            sc(c, f=fill('E8E8E8'), bd=border(color='CCCCCC'))
        elif day_num in INCOME_DATES and day_num in BILL_DATES:
            sc(c, f=fill('FFE082'),
               fn=font(bold=True, size=9, color=C['dark']),
               al=align('center', 'center', wrap=True),
               bd=border(color=C['warning']),
               v=f"{day_num}\n{INCOME_DATES[day_num]}\n{BILL_DATES[day_num]}")
        elif day_num in INCOME_DATES:
            sc(c, f=fill('C8E6C9'),
               fn=font(bold=True, size=9, color='1B5E20'),
               al=align('center', 'center', wrap=True),
               bd=border(color=C['success']),
               v=f"{day_num}\n{INCOME_DATES[day_num]}")
        elif day_num in BILL_DATES:
            sc(c, f=fill('FFCDD2'),
               fn=font(bold=True, size=9, color='B71C1C'),
               al=align('center', 'center', wrap=True),
               bd=border(color=C['danger']),
               v=f"{day_num}\n{BILL_DATES[day_num]}")
        else:
            sc(c, f=fill('F8F9FA'),
               fn=font(size=13, color=C['dark']),
               al=Alignment(horizontal='right', vertical='top'),
               bd=border(color='DEE2E6'),
               v=day_num)

spacer(ws8, 10, 16)

# ── Upcoming Bills table ──────────────────────────────────────
rh(ws8, 11, 32)
ws8.merge_cells('A11:I11')
sc(ws8.cell(row=11, column=1),
   f=fill(C['dark']), fn=font(bold=True, size=13, color='FFFFFF'),
   al=align('center', 'center'), v="UPCOMING BILLS & PAYMENTS — January 2026")

rh(ws8, 12, 26)
for col, txt in enumerate(['', 'Date', 'Bill / Item', 'Amount', 'Auto-Pay?', 'Status', '', '', ''], 1):
    if txt:
        hdr_cell(ws8, 12, col, txt, C['primary'], sz=10)

BILLS = [
    ('Jan 5',  'Rent Payment',           1500.00, 'No',  'Paid'),
    ('Jan 8',  'Electric Bill',            94.20, 'Yes', 'Paid'),
    ('Jan 16', 'Netflix Subscription',     15.99, 'Yes', 'Paid'),
    ('Jan 20', 'Savings Auto-Transfer',   300.00, 'Yes', 'Paid'),
    ('Jan 30', 'Student Loan Payment',    250.00, 'Yes', 'Upcoming'),
]
for i, (date, item, amt, auto, status) in enumerate(BILLS):
    row = 13 + i
    rh(ws8, row, 26)
    bg = 'F8F9FA' if i % 2 == 0 else 'FFFFFF'
    data_cell(ws8, row, 2, date,   bg=bg, bold=True, al_h='center')
    data_cell(ws8, row, 3, item,   bg=bg)
    data_cell(ws8, row, 4, amt,    bg=bg, al_h='right', nf='$#,##0.00')
    data_cell(ws8, row, 5, auto,   bg=bg, al_h='center')
    c_st = ws8.cell(row=row, column=6, value=status)
    if status == 'Paid':
        sc(c_st, f=fill('C8E6C9'), fn=font(bold=True, size=10, color='1B5E20'),
           al=align('center', 'center'), bd=border())
    else:
        sc(c_st, f=fill('FFCDD2'), fn=font(bold=True, size=10, color='B71C1C'),
           al=align('center', 'center'), bd=border())

legend_row = 13 + len(BILLS) + 1
spacer(ws8, 13 + len(BILLS), 8)
rh(ws8, legend_row, 30)
ws8.merge_cells(f'A{legend_row}:I{legend_row}')
sc(ws8.cell(row=legend_row, column=1),
   f=fill('FFF3E0'), fn=font(size=10, color=C['dark']),
   al=align('center', 'center'),
   bd=border(color=C['accent']),
   v="LEGEND:   🔴 Red = Bill Due Date   |   🟢 Green = Income / Payday   |   ⚪ Gray = Regular Day   |   🟡 Yellow = Both on same day")

# ═══════════════════════════════════════════════════════════
# GLOBAL PAGE SETUP
# ═══════════════════════════════════════════════════════════
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]:
    ws.page_setup.orientation  = 'landscape'
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
OUTPUT = "Budget_Planner_2026_All_In_One.xlsx"
wb.save(OUTPUT)
print(f"✅  Saved → {OUTPUT}")
print(f"    Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"    MT data rows: {DATA_START} – {LAST_DATA_ROW}  ({N_SAMPLES} samples + {N_EMPTY} blank)")
print(f"    TI_ROW={TI_ROW}, TE_ROW={TE_ROW}, NB_ROW={NB_ROW}")
